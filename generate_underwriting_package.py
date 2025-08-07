#!/usr/bin/env python3
"""
Generate professional underwriting package from extracted test data.
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from document_processor import DocumentProcessor
from underwriting_analyzer import UnderwritingAnalyzer
from underwriting_output import UnderwritingOutputGenerator
import logging

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def generate_underwriting_package():
    """Generate professional underwriting package from test data."""
    
    print("🚀 Generating Professional Underwriting Package...")
    print("=" * 60)
    
    # Path to test files
    test_dir = "uploads/2ed8d504-4f6a-4bd2-ab3b-8cd5c137a5cb"
    rent_roll_path = f"{test_dir}/rent_roll/RR_3350_Mount_Gilead_Rd_Atlanta_GA_30311.pdf"
    t12_path = f"{test_dir}/t12/T12_3350_Mount_Gilead_Rd_Atlanta_GA_30311.pdf"
    
    # Initialize processors
    processor = DocumentProcessor(debug=True)
    analyzer = UnderwritingAnalyzer(debug=True)
    output_generator = UnderwritingOutputGenerator(debug=True)
    
    # Extract data
    print("\n📊 Extracting data...")
    
    # Process rent roll
    rent_roll_results = processor.process_document(rent_roll_path)
    rent_roll_df = rent_roll_results['tables'][0] if rent_roll_results.get('tables') else None
    
    # Process T12
    t12_results = processor.process_document(t12_path)
    t12_df = t12_results['tables'][0] if t12_results.get('tables') else None
    
    if rent_roll_df is None or t12_df is None:
        print("❌ Failed to extract data from documents")
        return
    
    # Clean and prepare data
    print("\n🧹 Cleaning and preparing data...")
    
    # Clean rent roll
    clean_rent_roll = clean_rent_roll_data(rent_roll_df)
    
    # Clean T12
    clean_t12 = clean_t12_data(t12_df)
    
    # Generate underwriting summary
    print("\n📈 Generating underwriting summary...")
    underwriting_summary = generate_underwriting_summary(clean_rent_roll, clean_t12)
    
    # Create Excel package
    print("\n📊 Creating Excel package...")
    excel_path = create_excel_package(clean_rent_roll, clean_t12, underwriting_summary)
    
    # Create PDF package
    print("\n📄 Creating PDF package...")
    pdf_path = create_pdf_package(excel_path, clean_rent_roll, clean_t12, underwriting_summary)
    
    print(f"\n✅ Underwriting package generated successfully!")
    print(f"   - Excel: {excel_path}")
    print(f"   - PDF: {pdf_path}")
    
    return {
        'excel_path': excel_path,
        'pdf_path': pdf_path,
        'rent_roll': clean_rent_roll,
        't12': clean_t12,
        'summary': underwriting_summary
    }

def clean_rent_roll_data(rent_roll_df):
    """Clean rent roll data for underwriting package."""
    
    print("   🏠 Cleaning rent roll data...")
    
    # Rename columns for clarity
    column_mapping = {
        '0': 'Unit_Number',
        '1': 'Unit_ID',
        '2': 'Unit_Type',
        '3': 'Square_Feet',
        '4': 'Tenant_Name',
        '5': 'Current_Rent',
        '6': 'Water_Fees',
        '7': 'Pest_Trash_Fees',
        '8': 'Lease_Term',
        '9': 'Lease_Begin_Date',
        '10': 'Lease_End_Date'
    }
    
    # Rename columns
    rent_roll_df = rent_roll_df.rename(columns=column_mapping)
    
    # Clean data
    clean_data = []
    for idx, row in rent_roll_df.iterrows():
        if idx == 0:  # Skip header row
            continue
            
        # Extract unit information
        unit_number = str(row['Unit_Number']).strip()
        if not unit_number or unit_number == 'nan' or unit_number == 'Unit #':
            continue
            
        # Clean rent amount
        rent_str = str(row['Current_Rent']).replace('$', '').replace(',', '').strip()
        try:
            current_rent = float(rent_str) if rent_str and rent_str != 'nan' else 0
        except:
            current_rent = 0
            
        # Determine occupancy status
        tenant_name = str(row['Tenant_Name']).strip()
        status = 'Occupied' if tenant_name and tenant_name != 'nan' else 'Vacant'
        
        # Clean square footage
        sqft_str = str(row['Square_Feet']).replace(',', '').strip()
        try:
            square_feet = float(sqft_str) if sqft_str and sqft_str != 'nan' else 1187
        except:
            square_feet = 1187
            
        clean_data.append({
            'Unit_Number': unit_number,
            'Unit_Type': str(row['Unit_Type']).strip(),
            'Square_Feet': square_feet,
            'Current_Rent': current_rent,
            'Status': status,
            'Lease_End_Date': str(row['Lease_End_Date']).strip()
        })
    
    clean_df = pd.DataFrame(clean_data)
    
    # Calculate additional metrics
    clean_df['Rent_per_SqFt'] = clean_df['Current_Rent'] / clean_df['Square_Feet']
    clean_df['Annual_Rent'] = clean_df['Current_Rent'] * 12
    
    print(f"   ✅ Cleaned {len(clean_df)} units")
    print(f"   - Average rent: ${clean_df['Current_Rent'].mean():,.0f}")
    print(f"   - Average rent per sq ft: ${clean_df['Rent_per_SqFt'].mean():.2f}")
    print(f"   - Occupied units: {len(clean_df[clean_df['Status'] == 'Occupied'])}")
    print(f"   - Vacant units: {len(clean_df[clean_df['Status'] == 'Vacant'])}")
    
    return clean_df

def clean_t12_data(t12_df):
    """Clean T12 data for underwriting package."""
    
    print("   💰 Cleaning T12 data...")
    
    # Extract key financial metrics
    financial_data = {}
    
    for idx, row in t12_df.iterrows():
        row_text = str(row.iloc[0]).strip()
        
        # Extract key metrics
        if 'Gross Potential Rents' in row_text:
            financial_data['gross_potential_rents'] = extract_amount(row.iloc[-1])
        elif 'Loss to Lease' in row_text:
            financial_data['loss_to_lease'] = extract_amount(row.iloc[-1])
        elif 'Vacancy Loss' in row_text:
            financial_data['vacancy_loss'] = extract_amount(row.iloc[-1])
        elif 'TOTAL PROPERTY RENTAL INCOME' in row_text:
            financial_data['total_rental_income'] = extract_amount(row.iloc[-1])
        elif 'TOTAL OTHER INCOME' in row_text:
            financial_data['total_other_income'] = extract_amount(row.iloc[-1])
        elif 'TOTAL REVENUES' in row_text:
            financial_data['total_revenues'] = extract_amount(row.iloc[-1])
        elif 'Property Taxes' in row_text:
            financial_data['property_taxes'] = extract_amount(row.iloc[-1])
        elif 'Insurance Premiums' in row_text:
            financial_data['insurance'] = extract_amount(row.iloc[-1])
        elif 'TOTAL UTILITIES' in row_text:
            financial_data['utilities'] = extract_amount(row.iloc[-1])
        elif 'Maintenance/Repairs' in row_text:
            financial_data['maintenance_repairs'] = extract_amount(row.iloc[-1])
        elif 'Management Fees' in row_text:
            financial_data['management_fees'] = extract_amount(row.iloc[-1])
        elif 'TOTAL OPERATING EXPENSES' in row_text:
            financial_data['total_operating_expenses'] = extract_amount(row.iloc[-1])
        elif 'NET OPERATING INCOME' in row_text:
            financial_data['net_operating_income'] = extract_amount(row.iloc[-1])
    
    print(f"   ✅ Extracted key financial metrics")
    print(f"   - Gross Potential Rents: ${financial_data.get('gross_potential_rents', 0):,.0f}")
    print(f"   - Net Operating Income: ${financial_data.get('net_operating_income', 0):,.0f}")
    
    return financial_data

def extract_amount(amount_str):
    """Extract numeric amount from string."""
    if pd.isna(amount_str):
        return 0
    
    amount_str = str(amount_str).replace('$', '').replace(',', '').replace('(', '').replace(')', '').strip()
    try:
        return float(amount_str) if amount_str and amount_str != 'nan' else 0
    except:
        return 0

def generate_underwriting_summary(rent_roll_df, t12_data):
    """Generate underwriting summary with rulebook compliance."""
    
    print("   📊 Generating underwriting summary...")
    
    # Calculate key metrics
    total_units = len(rent_roll_df)
    occupied_units = len(rent_roll_df[rent_roll_df['Status'] == 'Occupied'])
    vacant_units = total_units - occupied_units
    
    # Income calculations
    gross_potential_rents = t12_data.get('gross_potential_rents', 0)
    total_rental_income = t12_data.get('total_rental_income', 0)
    total_other_income = t12_data.get('total_other_income', 0)
    
    # Expense calculations with rulebook adjustments
    property_taxes = t12_data.get('property_taxes', 0) * 1.075  # 7.5% increase for refinance
    insurance = t12_data.get('insurance', 0) * 1.05  # 5% increase
    utilities = t12_data.get('utilities', 0) * 1.02  # 2% increase
    maintenance_repairs = t12_data.get('maintenance_repairs', 0)
    management_fees = t12_data.get('management_fees', 0)
    
    # Apply rulebook adjustments
    # R&M minimum (assuming 25-year property age)
    rm_minimum = 700 * total_units  # $700/unit for 20-30 year property
    maintenance_repairs = max(maintenance_repairs, rm_minimum)
    
    # Management fee tiers
    if gross_potential_rents <= 500000:
        mgmt_rate = 0.05
    elif gross_potential_rents <= 750000:
        mgmt_rate = 0.045
    elif gross_potential_rents <= 1000000:
        mgmt_rate = 0.04
    elif gross_potential_rents <= 1500000:
        mgmt_rate = 0.035
    else:
        mgmt_rate = 0.03
    
    management_fees = gross_potential_rents * mgmt_rate
    
    # Replacement reserves
    replacement_reserves = 250 * total_units  # $250/unit
    
    # Calculate NOI
    total_expenses = property_taxes + insurance + utilities + maintenance_repairs + management_fees + replacement_reserves
    net_operating_income = total_rental_income + total_other_income - total_expenses
    
    # Create summary
    summary = {
        'property_info': {
            'name': 'Bolden Heights Apartments',
            'address': '3350 Mount Gilead Road, Atlanta, GA 30311',
            'total_units': total_units,
            'occupied_units': occupied_units,
            'vacant_units': vacant_units
        },
        'income_analysis': {
            'gross_potential_rents': gross_potential_rents,
            'total_rental_income': total_rental_income,
            'total_other_income': total_other_income,
            'effective_gross_income': total_rental_income + total_other_income
        },
        'expense_analysis': {
            'property_taxes': property_taxes,
            'insurance': insurance,
            'utilities': utilities,
            'maintenance_repairs': maintenance_repairs,
            'management_fees': management_fees,
            'replacement_reserves': replacement_reserves,
            'total_expenses': total_expenses
        },
        'noi_analysis': {
            'net_operating_income': net_operating_income,
            'expense_ratio': total_expenses / (total_rental_income + total_other_income) if (total_rental_income + total_other_income) > 0 else 0
        }
    }
    
    print(f"   ✅ Underwriting summary generated")
    print(f"   - NOI: ${net_operating_income:,.0f}")
    print(f"   - Expense ratio: {summary['noi_analysis']['expense_ratio']:.1%}")
    
    return summary

def create_excel_package(rent_roll_df, t12_data, summary):
    """Create Excel package with all required tabs."""
    
    print("   📊 Creating Excel package...")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # 1. Clean Rent Roll Tab
    ws_rent_roll = wb.create_sheet("Clean Rent Roll")
    
    # Add headers
    headers = ['Unit Number', 'Unit Type', 'Square Feet', 'Current Rent', 'Status', 'Lease End Date', 'Rent per SqFt', 'Annual Rent']
    for col, header in enumerate(headers, 1):
        cell = ws_rent_roll.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Add data
    for row_idx, row in rent_roll_df.iterrows():
        for col_idx, value in enumerate(row.values, 1):
            ws_rent_roll.cell(row=row_idx+2, column=col_idx, value=value)
    
    # 2. Clean T12 Tab
    ws_t12 = wb.create_sheet("Clean T12")
    
    # Add T12 data
    t12_headers = ['Line Item', 'Amount', 'Notes']
    for col, header in enumerate(t12_headers, 1):
        cell = ws_t12.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    t12_data_rows = [
        ['Gross Potential Rents', f"${t12_data.get('gross_potential_rents', 0):,.0f}", ''],
        ['Total Rental Income', f"${t12_data.get('total_rental_income', 0):,.0f}", ''],
        ['Total Other Income', f"${t12_data.get('total_other_income', 0):,.0f}", ''],
        ['Effective Gross Income', f"${t12_data.get('total_rental_income', 0) + t12_data.get('total_other_income', 0):,.0f}", ''],
        ['', '', ''],
        ['OPERATING EXPENSES', '', ''],
        ['Property Taxes', f"${summary['expense_analysis']['property_taxes']:,.0f}", 'Adjusted +7.5% for refinance'],
        ['Insurance', f"${summary['expense_analysis']['insurance']:,.0f}", 'Adjusted +5%'],
        ['Utilities', f"${summary['expense_analysis']['utilities']:,.0f}", 'Adjusted +2%'],
        ['Maintenance & Repairs', f"${summary['expense_analysis']['maintenance_repairs']:,.0f}", 'Age-based minimum applied'],
        ['Management Fees', f"${summary['expense_analysis']['management_fees']:,.0f}", 'Tier-based calculation'],
        ['Replacement Reserves', f"${summary['expense_analysis']['replacement_reserves']:,.0f}", '$250/unit'],
        ['Total Operating Expenses', f"${summary['expense_analysis']['total_expenses']:,.0f}", ''],
        ['', '', ''],
        ['NET OPERATING INCOME', f"${summary['noi_analysis']['net_operating_income']:,.0f}", '']
    ]
    
    for row_idx, row_data in enumerate(t12_data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_t12.cell(row=row_idx, column=col_idx, value=value)
            if row_data[0] in ['OPERATING EXPENSES', 'NET OPERATING INCOME']:
                cell.font = Font(bold=True)
    
    # 3. Underwriting Summary Tab
    ws_summary = wb.create_sheet("Underwriting Summary")
    
    # Add summary data
    summary_headers = ['Line Item', '$ Amount', '% of EGI', 'Notes']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    egi = summary['income_analysis']['effective_gross_income']
    summary_data_rows = [
        ['Rental Income', f"${summary['income_analysis']['total_rental_income']:,.0f}", f"{summary['income_analysis']['total_rental_income']/egi*100:.1f}%", ''],
        ['Other Income', f"${summary['income_analysis']['total_other_income']:,.0f}", f"{summary['income_analysis']['total_other_income']/egi*100:.1f}%", ''],
        ['Effective Gross Income', f"${egi:,.0f}", '100.0%', ''],
        ['', '', '', ''],
        ['OPERATING EXPENSES', '', '', ''],
        ['Property Taxes', f"${summary['expense_analysis']['property_taxes']:,.0f}", f"{summary['expense_analysis']['property_taxes']/egi*100:.1f}%", 'Adjusted +7.5%'],
        ['Insurance', f"${summary['expense_analysis']['insurance']:,.0f}", f"{summary['expense_analysis']['insurance']/egi*100:.1f}%", 'Adjusted +5%'],
        ['Utilities', f"${summary['expense_analysis']['utilities']:,.0f}", f"{summary['expense_analysis']['utilities']/egi*100:.1f}%", 'Adjusted +2%'],
        ['Maintenance & Repairs', f"${summary['expense_analysis']['maintenance_repairs']:,.0f}", f"{summary['expense_analysis']['maintenance_repairs']/egi*100:.1f}%", 'Age-based minimum'],
        ['Management Fees', f"${summary['expense_analysis']['management_fees']:,.0f}", f"{summary['expense_analysis']['management_fees']/egi*100:.1f}%", 'Tier-based calculation'],
        ['Replacement Reserves', f"${summary['expense_analysis']['replacement_reserves']:,.0f}", f"{summary['expense_analysis']['replacement_reserves']/egi*100:.1f}%", '$250/unit'],
        ['Total Operating Expenses', f"${summary['expense_analysis']['total_expenses']:,.0f}", f"{summary['expense_analysis']['total_expenses']/egi*100:.1f}%", ''],
        ['', '', '', ''],
        ['NET OPERATING INCOME', f"${summary['noi_analysis']['net_operating_income']:,.0f}", f"{summary['noi_analysis']['net_operating_income']/egi*100:.1f}%", '']
    ]
    
    for row_idx, row_data in enumerate(summary_data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_data[0] in ['OPERATING EXPENSES', 'NET OPERATING INCOME']:
                cell.font = Font(bold=True)
    
    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = f"outputs/Bolden_Heights_Underwriting_Package_{timestamp}.xlsx"
    os.makedirs("outputs", exist_ok=True)
    wb.save(excel_path)
    
    print(f"   ✅ Excel package created: {excel_path}")
    return excel_path

def create_pdf_package(excel_path, rent_roll_df, t12_data, summary):
    """Create PDF package from Excel file."""
    
    print("   📄 Creating PDF package...")
    
    # For now, we'll create a simple PDF summary
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    pdf_path = f"outputs/Bolden_Heights_Underwriting_Package_{timestamp}.pdf"
    
    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("Bolden Heights Apartments - Underwriting Package", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Property Information
    story.append(Paragraph("Property Information", styles['Heading2']))
    prop_info = [
        ['Property Name:', summary['property_info']['name']],
        ['Address:', summary['property_info']['address']],
        ['Total Units:', str(summary['property_info']['total_units'])],
        ['Occupied Units:', str(summary['property_info']['occupied_units'])],
        ['Vacant Units:', str(summary['property_info']['vacant_units'])]
    ]
    
    prop_table = Table(prop_info, colWidths=[2*72, 4*72])
    prop_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(prop_table)
    story.append(Spacer(1, 20))
    
    # Financial Summary
    story.append(Paragraph("Financial Summary", styles['Heading2']))
    financial_data = [
        ['Gross Potential Rents:', f"${summary['income_analysis']['gross_potential_rents']:,.0f}"],
        ['Effective Gross Income:', f"${summary['income_analysis']['effective_gross_income']:,.0f}"],
        ['Total Operating Expenses:', f"${summary['expense_analysis']['total_expenses']:,.0f}"],
        ['Net Operating Income:', f"${summary['noi_analysis']['net_operating_income']:,.0f}"],
        ['Expense Ratio:', f"{summary['noi_analysis']['expense_ratio']:.1%}"]
    ]
    
    financial_table = Table(financial_data, colWidths=[2*72, 4*72])
    financial_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(financial_table)
    
    doc.build(story)
    
    print(f"   ✅ PDF package created: {pdf_path}")
    return pdf_path

if __name__ == "__main__":
    # Generate underwriting package
    results = generate_underwriting_package()
    
    if results:
        print(f"\n🎯 Underwriting package generation complete!")
        print(f"   - Excel file: {results['excel_path']}")
        print(f"   - PDF file: {results['pdf_path']}")
        print(f"   - Total units analyzed: {len(results['rent_roll'])}")
        print(f"   - NOI: ${results['summary']['noi_analysis']['net_operating_income']:,.0f}")
    else:
        print(f"\n❌ Failed to generate underwriting package") 