#!/usr/bin/env python3
"""
Analyze the UW template structure from Hardwell_UW_Example deal 1.xlsx
"""

import pandas as pd
import openpyxl
import os

def analyze_uw_template():
    """Analyze the UW template structure."""
    
    excel_path = "../Hardwell_UW_Example deal 1.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        return
    
    print("🔍 Analyzing UW Template Structure...")
    print("=" * 80)
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        print(f"📊 Worksheet Names:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"   {i}. {sheet_name}")
        
        print("\n" + "=" * 80)
        
        # Focus on UW tab if it exists
        uw_sheets = [name for name in wb.sheetnames if 'uw' in name.lower() or 'underwriting' in name.lower()]
        
        if uw_sheets:
            print(f"\n🎯 Found UW-related sheets: {uw_sheets}")
            
            for uw_sheet in uw_sheets:
                print(f"\n📋 Analyzing UW Sheet: '{uw_sheet}'")
                print("-" * 60)
                
                ws = wb[uw_sheet]
                
                # Get sheet dimensions
                max_row = ws.max_row
                max_col = ws.max_column
                print(f"   Dimensions: {max_row} rows × {max_col} columns")
                
                # Show detailed structure
                print(f"   📊 UW Template Structure (first 50 rows):")
                
                for row in range(1, min(51, max_row + 1)):
                    row_data = []
                    has_data = False
                    
                    for col in range(1, min(15, max_col + 1)):  # First 15 columns
                        cell = ws.cell(row=row, column=col)
                        value = cell.value
                        
                        if value is not None:
                            has_data = True
                            if isinstance(value, (int, float)):
                                if abs(value) > 1000:
                                    row_data.append(f"${value:,.0f}")
                                elif abs(value) > 1:
                                    row_data.append(f"{value:.2f}")
                                else:
                                    row_data.append(f"{value:.4f}")
                            else:
                                # Truncate long strings
                                str_val = str(value)[:40]
                                row_data.append(str_val)
                        else:
                            row_data.append("")
                    
                    if has_data:
                        # Show row with column indicators
                        col_headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
                        formatted_row = []
                        for i, data in enumerate(row_data):
                            if data.strip():
                                formatted_row.append(f"{col_headers[i]}:{data}")
                        
                        if formatted_row:
                            print(f"   Row {row:2d}: {' | '.join(formatted_row)}")
        
        else:
            print("\n⚠️ No UW-specific sheets found. Analyzing all sheets...")
            
            # Analyze each sheet for UW-related content
            for sheet_name in wb.sheetnames:
                print(f"\n📋 Analyzing Sheet: '{sheet_name}'")
                print("-" * 50)
                
                ws = wb[sheet_name]
                
                # Look for underwriting-related content
                uw_content_found = False
                for row in range(1, min(30, ws.max_row + 1)):
                    for col in range(1, min(10, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            if any(keyword in cell_value.lower() for keyword in 
                                   ['underwriting', 'noi', 'cap rate', 'dscr', 'ltv', 'debt yield', 
                                    'gross income', 'operating expense', 'cash flow']):
                                if not uw_content_found:
                                    print(f"   🎯 Found UW content in {sheet_name}:")
                                    uw_content_found = True
                                print(f"     Row {row}, Col {col}: {cell_value}")
                
                if not uw_content_found:
                    print(f"   ℹ️ No obvious UW content found in {sheet_name}")
        
        print("\n" + "=" * 80)
        print("✅ UW Template analysis complete!")
        
    except Exception as e:
        print(f"❌ Error analyzing template: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_uw_template()