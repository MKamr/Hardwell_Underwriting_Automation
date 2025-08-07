#!/usr/bin/env python3
"""
Template-Based Underwriting Generator
Uses the existing Excel template and fills it with rulebook-compliant extracted data
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from document_processor import DocumentProcessor
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

class TemplateBasedGenerator:
    """Generator that uses existing Excel template and fills with extracted data."""
    
    def __init__(self, template_path, debug=False):
        self.template_path = template_path
        self.debug = debug
        self.processor = DocumentProcessor(debug=debug)
        
        # Rulebook configuration (same as before)
        self.rulebook_config = {
            'vacancy_minimum': 0.05,  # 5% minimum
            'property_tax_adjustments': {
                'refinance': 1.075,  # +7.5% for refinance
                'acquisition': 1.0
            },
            'insurance_adjustment': 1.05,  # +5%
            'utilities_adjustment': 1.02,  # +2%
            'rm_minimums_by_age': {
                (0, 10): 500, (10, 20): 600, (20, 30): 700,
                (30, 40): 800, (40, 50): 900, (50, 100): 1000
            },
            'rm_cap': 1500,  # Cap at $1,500/unit
            'management_fee_tiers': [
                (0, 500000, 0.05), (500000, 750000, 0.045),
                (750000, 1000000, 0.04), (1000000, 1500000, 0.035),
                (1500000, 2000000, 0.03), (2000000, float('inf'), 0.025)
            ],
            'professional_fees': {'minimum': 1000, 'maximum_per_unit': 400},
            'replacement_reserves': 250,  # $250/unit
            'minimum_expense_ratio': 0.28  # 28% minimum
        }
    
    def generate_from_template(self, rent_roll_path, t12_path, property_info):
        """Generate underwriting package using existing template."""
        
        print("🚀 Generating Template-Based Underwriting Package...")
        print("=" * 80)
        
        # Step 1: Extract and process data (same as rulebook compliant)
        print("📊 Extracting data using rulebook compliance...")
        rent_roll_data = self._extract_rent_roll(rent_roll_path)
        t12_data = self._extract_t12(t12_path)
        
        # Step 2: Apply rulebook rules
        print("💰 Applying rulebook rules...")
        income_analysis = self._apply_income_rules(rent_roll_data, t12_data)
        expense_analysis = self._apply_expense_rules(t12_data, income_analysis, property_info)
        noi_analysis = self._calculate_noi_and_validate(income_analysis, expense_analysis)
        
        # Step 3: Load template and fill with data
        print("📋 Loading template and filling with extracted data...")
        filled_excel_path = self._fill_template_with_data(
            rent_roll_data, t12_data, income_analysis, expense_analysis, noi_analysis, property_info
        )
        
        print(f"\n✅ Template-based package generated successfully!")
        print(f"   📊 Excel: {filled_excel_path}")
        print(f"   💰 NOI: ${noi_analysis['net_operating_income']:,.0f}")
        print(f"   📈 Expense Ratio: {expense_analysis['expense_ratio']:.1%}")
        
        return {
            'excel_path': filled_excel_path,
            'rent_roll_data': rent_roll_data,
            'income_analysis': income_analysis,
            'expense_analysis': expense_analysis,
            'noi_analysis': noi_analysis
        }
    
    def _extract_rent_roll(self, rent_roll_path):
        """Extract rent roll data."""
        results = self.processor.process_document(rent_roll_path)
        rent_roll_df = results['tables'][0] if results.get('tables') else None
        
        if rent_roll_df is None:
            raise ValueError("Failed to extract rent roll data")
        
        # Clean data
        cleaned_data = []
        for idx, row in rent_roll_df.iterrows():
            if idx == 0:  # Skip header
                continue
                
            unit_number = str(row.iloc[0]).strip()
            if not unit_number or unit_number == 'nan' or 'Unit' in unit_number:
                continue
            
            unit_type = str(row.iloc[2]).strip()
            square_feet = self._safe_float(row.iloc[3], 1187)
            current_rent = self._safe_float(row.iloc[5], 0)
            tenant_name = str(row.iloc[4]).strip()
            is_occupied = tenant_name and tenant_name != 'nan'
            lease_end_date = str(row.iloc[10]).strip() if len(row) > 10 else ''
            
            cleaned_data.append({
                'Unit_Number': unit_number,
                'Unit_ID': str(row.iloc[1]).strip(),
                'Unit_Type': unit_type,
                'Square_Feet': square_feet,
                'Current_Rent': current_rent,
                'Is_Occupied': is_occupied,
                'Tenant_Name': tenant_name if is_occupied else '',
                'Lease_End_Date': lease_end_date,
                'Water_Fees': self._safe_float(row.iloc[6], 65),
                'Pest_Trash_Fees': self._safe_float(row.iloc[7], 15),
                'Lease_Term': str(row.iloc[8]).strip() if len(row) > 8 else '12-Month'
            })
        
        return pd.DataFrame(cleaned_data)
    
    def _extract_t12(self, t12_path):
        """Extract T12 data."""
        results = self.processor.process_document(t12_path)
        t12_df = results['tables'][0] if results.get('tables') else None
        
        if t12_df is None:
            raise ValueError("Failed to extract T12 data")
        
        # Extract monthly and annual data
        financial_data = {}
        monthly_data = {}
        
        for idx, row in t12_df.iterrows():
            row_text = str(row.iloc[0]).strip().upper()
            
            # Get annual total (last column)
            annual_amount = self._safe_float(row.iloc[-1], 0)
            
            # Get monthly data (columns 1-12)
            monthly_amounts = []
            for col in range(1, min(13, len(row))):
                monthly_amounts.append(self._safe_float(row.iloc[col], 0))
            
            # Map line items
            if 'GROSS POTENTIAL RENT' in row_text:
                financial_data['gross_potential_rents'] = annual_amount
                monthly_data['gross_potential_rents'] = monthly_amounts
            elif 'LOSS TO LEASE' in row_text:
                financial_data['loss_to_lease'] = annual_amount
                monthly_data['loss_to_lease'] = monthly_amounts
            elif 'VACANCY LOSS' in row_text:
                financial_data['vacancy_loss'] = annual_amount
                monthly_data['vacancy_loss'] = monthly_amounts
            elif 'TOTAL PROPERTY RENTAL INCOME' in row_text:
                financial_data['rental_income'] = annual_amount
                monthly_data['rental_income'] = monthly_amounts
            elif 'TOTAL OTHER INCOME' in row_text:
                financial_data['other_income'] = annual_amount
                monthly_data['other_income'] = monthly_amounts
            elif 'PROPERTY TAXES' in row_text:
                financial_data['property_taxes'] = annual_amount
                monthly_data['property_taxes'] = monthly_amounts
            elif 'INSURANCE' in row_text and 'TOTAL' in row_text:
                financial_data['insurance'] = annual_amount
                monthly_data['insurance'] = monthly_amounts
            elif 'TOTAL UTILITIES' in row_text:
                financial_data['utilities'] = annual_amount
                monthly_data['utilities'] = monthly_amounts
            elif 'MAINTENANCE' in row_text and 'REPAIR' in row_text:
                financial_data['maintenance_repairs'] = annual_amount
                monthly_data['maintenance_repairs'] = monthly_amounts
            elif 'MANAGEMENT FEE' in row_text:
                financial_data['management_fees'] = annual_amount
                monthly_data['management_fees'] = monthly_amounts
            elif 'NET OPERATING INCOME' in row_text:
                financial_data['net_operating_income'] = annual_amount
                monthly_data['net_operating_income'] = monthly_amounts
                break  # Stop at NOI per rulebook
        
        financial_data['monthly_data'] = monthly_data
        return financial_data
    
    def _apply_income_rules(self, rent_roll_data, t12_data):
        """Apply income rules per rulebook."""
        occupied_units = rent_roll_data[rent_roll_data['Is_Occupied'] == True]
        vacant_units = rent_roll_data[rent_roll_data['Is_Occupied'] == False]
        
        # Calculate rental income
        occupied_rental_income = occupied_units['Current_Rent'].sum() * 12
        
        # Vacant unit income using average rent of unit type
        vacant_rental_income = 0
        for unit_type in vacant_units['Unit_Type'].unique():
            type_occupied = occupied_units[occupied_units['Unit_Type'] == unit_type]
            if len(type_occupied) > 0:
                avg_rent = type_occupied['Current_Rent'].mean()
                vacant_count = len(vacant_units[vacant_units['Unit_Type'] == unit_type])
                vacant_rental_income += avg_rent * vacant_count * 12
        
        total_rental_income = occupied_rental_income + vacant_rental_income
        other_income = t12_data.get('other_income', 0)
        gross_potential_income = total_rental_income + other_income
        
        return {
            'occupied_rental_income': occupied_rental_income,
            'vacant_rental_income': vacant_rental_income,
            'total_rental_income': total_rental_income,
            'other_income': other_income,
            'gross_potential_income': gross_potential_income
        }
    
    def _apply_expense_rules(self, t12_data, income_analysis, property_info):
        """Apply expense rules per rulebook."""
        gross_potential_income = income_analysis['gross_potential_income']
        total_units = getattr(property_info, 'total_units', 86)
        property_age = getattr(property_info, 'property_age', 25)
        transaction_type = getattr(property_info, 'transaction_type', 'refinance')
        
        # Vacancy: 5% minimum or actual (whichever higher)
        actual_vacancy = gross_potential_income - income_analysis['total_rental_income']
        minimum_vacancy = gross_potential_income * self.rulebook_config['vacancy_minimum']
        vacancy_loss = max(actual_vacancy, minimum_vacancy)
        
        effective_gross_income = gross_potential_income - vacancy_loss
        
        # Apply rulebook adjustments
        actual_property_taxes = t12_data.get('property_taxes', 0)
        if transaction_type == 'refinance':
            property_taxes = actual_property_taxes * self.rulebook_config['property_tax_adjustments']['refinance']
        else:
            property_taxes = actual_property_taxes
        
        actual_insurance = t12_data.get('insurance', 0)
        insurance = actual_insurance * self.rulebook_config['insurance_adjustment']
        
        actual_utilities = t12_data.get('utilities', 0)
        utilities = actual_utilities * self.rulebook_config['utilities_adjustment']
        
        # R&M with age-based minimums
        actual_rm = t12_data.get('maintenance_repairs', 0)
        rm_minimum = self._calculate_rm_minimum(total_units, property_age)
        maintenance_repairs = max(actual_rm, rm_minimum)
        rm_cap = total_units * self.rulebook_config['rm_cap']
        if maintenance_repairs > rm_cap:
            maintenance_repairs = rm_cap
        
        # Management fees with tiers
        management_fees = self._calculate_management_fees(gross_potential_income)
        
        # Professional fees
        prof_fees_min = self.rulebook_config['professional_fees']['minimum']
        prof_fees_max = total_units * self.rulebook_config['professional_fees']['maximum_per_unit']
        professional_fees = max(prof_fees_min, min(5000, prof_fees_max))
        
        # Replacement reserves
        replacement_reserves = total_units * self.rulebook_config['replacement_reserves']
        
        # Total expenses
        total_expenses = (property_taxes + insurance + utilities + maintenance_repairs + 
                         management_fees + professional_fees + replacement_reserves)
        
        # Enforce 28% minimum
        minimum_expenses = effective_gross_income * self.rulebook_config['minimum_expense_ratio']
        if total_expenses < minimum_expenses:
            shortage = minimum_expenses - total_expenses
            professional_fees += shortage
            total_expenses = minimum_expenses
        
        expense_ratio = total_expenses / effective_gross_income if effective_gross_income > 0 else 0
        
        return {
            'vacancy_loss': vacancy_loss,
            'effective_gross_income': effective_gross_income,
            'property_taxes': property_taxes,
            'insurance': insurance,
            'utilities': utilities,
            'maintenance_repairs': maintenance_repairs,
            'management_fees': management_fees,
            'professional_fees': professional_fees,
            'replacement_reserves': replacement_reserves,
            'total_expenses': total_expenses,
            'expense_ratio': expense_ratio
        }
    
    def _calculate_noi_and_validate(self, income_analysis, expense_analysis):
        """Calculate NOI."""
        net_operating_income = expense_analysis['effective_gross_income'] - expense_analysis['total_expenses']
        
        return {
            'net_operating_income': net_operating_income,
            'gross_potential_income': income_analysis['gross_potential_income'],
            'effective_gross_income': expense_analysis['effective_gross_income'],
            'total_expenses': expense_analysis['total_expenses'],
            'expense_ratio': expense_analysis['expense_ratio']
        }
    
    def _fill_template_with_data(self, rent_roll_data, t12_data, income_analysis, expense_analysis, noi_analysis, property_info):
        """Fill the existing template with extracted and processed data."""
        
        # Load the template
        wb = load_workbook(self.template_path)
        
        # Fill Rent Roll sheet
        self._fill_rent_roll_sheet(wb, rent_roll_data, property_info)
        
        # Fill T-12 sheet
        self._fill_t12_sheet(wb, t12_data, income_analysis, expense_analysis, noi_analysis)
        
        # Save the filled template
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = f"outputs/Filled_Template_Package_{timestamp}.xlsx"
        os.makedirs("outputs", exist_ok=True)
        wb.save(output_path)
        
        return output_path
    
    def _fill_rent_roll_sheet(self, wb, rent_roll_data, property_info):
        """Fill the Rent Roll sheet with cleaned data."""
        
        if 'Rent Roll' not in wb.sheetnames:
            print("⚠️ 'Rent Roll' sheet not found in template")
            return
        
        ws = wb['Rent Roll']
        
        # Update property information at the top
        ws['A1'] = getattr(property_info, 'property_name', 'Bolden Heights Apartments')
        ws['A2'] = getattr(property_info, 'property_address', '3350 Mount Gilead Road Atlanta')
        ws['A3'] = datetime.now().strftime('%B %d, %Y')
        
        # Update unit mix summary (rows 8-11)
        unit_types = rent_roll_data['Unit_Type'].value_counts()
        
        row = 8
        for unit_type, count in unit_types.items():
            type_data = rent_roll_data[rent_roll_data['Unit_Type'] == unit_type]
            avg_sqft = type_data['Square_Feet'].mean()
            avg_rent = type_data[type_data['Is_Occupied']]['Current_Rent'].mean()
            
            ws[f'C{row}'] = count  # Units
            ws[f'D{row}'] = unit_type  # Type
            ws[f'E{row}'] = f"${avg_sqft:,.0f}"  # Unit Square Feet
            ws[f'F{row}'] = f"${count * avg_sqft:,.0f}"  # Total SF
            ws[f'G{row}'] = f"${avg_rent * 1.08:,.0f}"  # Market Rent (8% above current)
            ws[f'H{row}'] = f"${count * avg_rent * 1.08:,.0f}"  # Monthly Market
            ws[f'I{row}'] = f"${count * avg_rent * 1.08 * 12:,.0f}"  # Annual Market
            
            row += 1
        
        # Fill detailed rent roll starting at row 14
        start_row = 14
        for idx, unit in rent_roll_data.iterrows():
            row = start_row + idx
            
            # Ensure we don't exceed Excel limits
            if row > 1000:
                break
            
            ws[f'A{row}'] = unit['Unit_Number']
            ws[f'B{row}'] = unit['Unit_ID']
            ws[f'C{row}'] = unit['Unit_Type']
            ws[f'D{row}'] = f"${unit['Square_Feet']:,.0f}"
            ws[f'E{row}'] = unit['Tenant_Name'] if unit['Is_Occupied'] else 'VACANT'
            ws[f'F{row}'] = f"${unit['Current_Rent']:,.0f}" if unit['Current_Rent'] > 0 else 'VACANT'
            ws[f'G{row}'] = unit['Water_Fees'] if unit['Is_Occupied'] else ''
            ws[f'H{row}'] = unit['Pest_Trash_Fees'] if unit['Is_Occupied'] else ''
            ws[f'I{row}'] = unit['Lease_Term'] if unit['Is_Occupied'] else ''
            
            # Format vacant units differently
            if not unit['Is_Occupied']:
                for col in ['E', 'F', 'G', 'H', 'I']:
                    cell = ws[f'{col}{row}']
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        print("   ✅ Rent Roll sheet updated")
    
    def _fill_t12_sheet(self, wb, t12_data, income_analysis, expense_analysis, noi_analysis):
        """Fill the T-12 sheet with financial data."""
        
        if 'T-12' not in wb.sheetnames:
            print("⚠️ 'T-12' sheet not found in template")
            return
        
        ws = wb['T-12']
        
        # Update header
        ws['A1'] = "Bolden Heights Apartments - Trailing 12 Month Operating Financials (UNDERWRITER ADJUSTED)"
        ws['A2'] = "3350 Mount Gilead Road Atlanta, GA 30311"
        
        # Fill key financial rows with rulebook-adjusted values
        # We'll update the annual totals (last column) with our calculated values
        
        # Find and update key rows
        for row in range(1, 60):  # Search through first 60 rows
            cell_value = ws[f'A{row}'].value
            if cell_value:
                cell_text = str(cell_value).upper()
                
                # Update annual totals with our rulebook-compliant calculations
                if 'GROSS POTENTIAL RENT' in cell_text:
                    ws[f'O{row}'] = f"${income_analysis['gross_potential_income']:,.0f}"
                elif 'VACANCY LOSS' in cell_text:
                    ws[f'O{row}'] = f"${-expense_analysis['vacancy_loss']:,.0f}"
                elif 'TOTAL PROPERTY RENTAL INCOME' in cell_text:
                    ws[f'O{row}'] = f"${income_analysis['total_rental_income']:,.0f}"
                elif 'TOTAL OTHER INCOME' in cell_text:
                    ws[f'O{row}'] = f"${income_analysis['other_income']:,.0f}"
                elif 'PROPERTY TAXES' in cell_text:
                    ws[f'O{row}'] = f"${expense_analysis['property_taxes']:,.0f}"
                    # Add note about adjustment
                    ws[f'P{row}'] = "Adjusted +7.5% for refinance per rulebook"
                elif 'INSURANCE' in cell_text and 'TOTAL' in cell_text:
                    ws[f'O{row}'] = f"${expense_analysis['insurance']:,.0f}"
                    ws[f'P{row}'] = "Adjusted +5% per rulebook"
                elif 'TOTAL UTILITIES' in cell_text:
                    ws[f'O{row}'] = f"${expense_analysis['utilities']:,.0f}"
                    ws[f'P{row}'] = "Adjusted +2% per rulebook"
                elif 'MAINTENANCE' in cell_text and 'REPAIR' in cell_text:
                    ws[f'O{row}'] = f"${expense_analysis['maintenance_repairs']:,.0f}"
                    ws[f'P{row}'] = "Age-based minimum applied per rulebook"
                elif 'MANAGEMENT FEE' in cell_text:
                    ws[f'O{row}'] = f"${expense_analysis['management_fees']:,.0f}"
                    ws[f'P{row}'] = "Tier-based calculation per rulebook"
                elif 'NET OPERATING INCOME' in cell_text:
                    ws[f'O{row}'] = f"${noi_analysis['net_operating_income']:,.0f}"
                    ws[f'P{row}'] = "Calculated with rulebook adjustments"
                    
                    # Add expense ratio
                    ws[f'A{row+1}'] = "Expense Ratio"
                    ws[f'O{row+1}'] = f"{expense_analysis['expense_ratio']:.1%}"
                    ws[f'P{row+1}'] = f"Meets {self.rulebook_config['minimum_expense_ratio']:.0%} minimum per rulebook"
                    break  # Stop at NOI per rulebook
        
        print("   ✅ T-12 sheet updated with rulebook adjustments")
    
    def _calculate_rm_minimum(self, total_units, property_age):
        """Calculate R&M minimum based on property age."""
        for (min_age, max_age), rate in self.rulebook_config['rm_minimums_by_age'].items():
            if min_age <= property_age < max_age:
                return total_units * rate
        return total_units * 1000
    
    def _calculate_management_fees(self, gross_potential_income):
        """Calculate management fees based on income tiers."""
        for min_income, max_income, rate in self.rulebook_config['management_fee_tiers']:
            if min_income <= gross_potential_income < max_income:
                return gross_potential_income * rate
        return gross_potential_income * 0.025
    
    def _safe_float(self, value, default=0):
        """Safely convert value to float."""
        if pd.isna(value):
            return default
        try:
            if isinstance(value, str):
                value = value.replace('$', '').replace(',', '').strip()
            return float(value) if value and str(value) != 'nan' else default
        except:
            return default

def main():
    """Main function to run the template-based generator."""
    
    class PropertyInfo:
        def __init__(self):
            self.property_name = "Bolden Heights Apartments"
            self.property_address = "3350 Mount Gilead Road, Atlanta, GA 30311"
            self.transaction_type = "refinance"
            self.property_age = 25
            self.total_units = 86
    
    property_info = PropertyInfo()
    
    # File paths
    template_path = "../Loan_Package_3350_Mount_Gilead_Rd_Atlanta_GA_30311_9_26_2024.xlsx"
    test_dir = "uploads/2ed8d504-4f6a-4bd2-ab3b-8cd5c137a5cb"
    rent_roll_path = f"{test_dir}/rent_roll/RR_3350_Mount_Gilead_Rd_Atlanta_GA_30311.pdf"
    t12_path = f"{test_dir}/t12/T12_3350_Mount_Gilead_Rd_Atlanta_GA_30311.pdf"
    
    # Generate template-based package
    generator = TemplateBasedGenerator(template_path, debug=True)
    
    try:
        results = generator.generate_from_template(rent_roll_path, t12_path, property_info)
        
        print(f"\n🎯 TEMPLATE-BASED package generated successfully!")
        print(f"   📊 Uses your existing template structure")
        print(f"   📋 Filled with rulebook-compliant data")
        print(f"   💰 NOI: ${results['noi_analysis']['net_operating_income']:,.0f}")
        print(f"   📈 Expense Ratio: {results['expense_analysis']['expense_ratio']:.1%}")
        
    except Exception as e:
        print(f"❌ Error generating package: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()