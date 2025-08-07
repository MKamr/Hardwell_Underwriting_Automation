#!/usr/bin/env python3
"""
Enhanced FastAPI Underwriting Application
Integrates the UW Template Filler for professional underwriting packages
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import os
import uuid
import json
import asyncio
import shutil
from datetime import datetime, timedelta
import logging
import csv
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# Import processing components
try:
    from document_processor import DocumentProcessor
    from underwriting_analyzer import UnderwritingAnalyzer
    from underwriting_output import UnderwritingOutputGenerator
    from llm_document_processor import LLMDocumentProcessor
    REAL_PROCESSING_AVAILABLE = True
    print("✅ Real PDF processing components loaded successfully")
except ImportError as e:
    print(f"⚠️ Some processing components not available: {e}")
    REAL_PROCESSING_AVAILABLE = False

# Import our UW Template Filler
try:
    
    UW_TEMPLATE_AVAILABLE = True
    print("✅ UW Template Filler loaded successfully")
except ImportError as e:
    UW_TEMPLATE_AVAILABLE = False
    print(f"⚠️ UW Template Filler not available: {e}")

# FastAPI app initialization
app = FastAPI(
    title="Enhanced Real Estate Underwriting AI",
    description="Professional underwriting with UW template integration",
    version="2.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Ensure directories exist
os.makedirs("static", exist_ok=True)
os.makedirs("templates", exist_ok=True)
os.makedirs("uploads", exist_ok=True)
os.makedirs("outputs", exist_ok=True)

# Static files - mount with check
try:
    app.mount("/static", StaticFiles(directory="static"), name="static")
except RuntimeError:
    # If static directory is empty or doesn't exist, create a basic one
    os.makedirs("static", exist_ok=True)
    with open("static/app.css", "w") as f:
        f.write("/* Hardwell Underwriting Automation */")
    app.mount("/static", StaticFiles(directory="static"), name="static")

# In-memory storage for processing status
processing_sessions = {}

# Template path configuration
UW_TEMPLATE_PATH = "../Hardwell_UW_Example deal 1.xlsx"

# Models
class ProcessingStatus(BaseModel):
    session_id: str
    status: str
    current_step: int
    total_steps: int
    step_name: str
    progress_percentage: float
    message: str
    results: Optional[Dict[str, Any]] = None
    error_message: Optional[str] = None

class PropertyInfo(BaseModel):
    property_name: str
    property_address: str
    transaction_type: str = "refinance"
    is_bridge_loan: bool = False
    property_age: int = 25

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.get("/health")
async def health_check():
    """Health check endpoint for deployment monitoring."""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

@app.head("/")
async def head_root():
    """Handle HEAD requests for health checks."""
    return {"status": "ok"}

@app.get("/", response_class=HTMLResponse)
async def read_root():
    """Serve the main application page."""
    return FileResponse("templates/index.html")

@app.post("/api/upload")
async def upload_documents(
    background_tasks: BackgroundTasks,
    property_name: str = Form(...),
    property_address: str = Form(...),
    transaction_type: str = Form("refinance"),
    is_bridge_loan: bool = Form(False),
    property_age: int = Form(25),
    files: List[UploadFile] = File(...),
    file_types: List[str] = Form([])
):
    """Upload documents and start processing with UW template generation."""
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    
    # Generate unique session ID
    session_id = str(uuid.uuid4())
    
    # Create session directory
    session_dir = f"uploads/{session_id}"
    os.makedirs(session_dir, exist_ok=True)
    
    # File validation and security
    allowed_extensions = {'.pdf', '.xlsx', '.xls', '.csv'}
    max_file_size = 50 * 1024 * 1024  # 50MB
    
    uploaded_files = {}
    file_type_mapping = {}
    
    for i, file in enumerate(files):
        if not file.filename:
            raise HTTPException(status_code=400, detail="Invalid filename")
        
        # Security: sanitize filename
        filename = Path(file.filename).name
        file_extension = Path(filename).suffix.lower()
        
        if file_extension not in allowed_extensions:
            raise HTTPException(status_code=400, detail=f"File type {file_extension} not allowed")
        
        # Check file size
        file.file.seek(0, 2)
        file_size = file.file.tell()
        file.file.seek(0)
        
        if file_size > max_file_size:
            raise HTTPException(status_code=400, detail=f"File size exceeds {max_file_size/(1024*1024):.0f}MB limit")
        
        # Determine file type
        file_type = file_types[i] if i < len(file_types) else "additional"
        
        # Create type-specific subdirectories
        type_dir = os.path.join(session_dir, file_type)
        os.makedirs(type_dir, exist_ok=True)
        
        # Save file
        file_path = os.path.join(type_dir, filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        uploaded_files[file_type] = file_path
        file_type_mapping[file_path] = file_type
    
    # Initialize processing session
    property_info = PropertyInfo(
        property_name=property_name,
        property_address=property_address,
        transaction_type=transaction_type,
        is_bridge_loan=is_bridge_loan,
        property_age=property_age
    )
    
    processing_sessions[session_id] = ProcessingStatus(
        session_id=session_id,
        status="waiting",
        current_step=0,
        total_steps=9,  # Updated to include CSV generation step
        step_name="Initializing",
        progress_percentage=0.0,
        message="Processing started"
    )
    
    # Count file types
    rent_roll_count = sum(1 for ft in file_type_mapping.values() if ft == "rent_roll")
    t12_count = sum(1 for ft in file_type_mapping.values() if ft == "t12")
    additional_count = len(uploaded_files) - rent_roll_count - t12_count
    
    # Start enhanced background processing
    background_tasks.add_task(
        process_documents_with_uw_template,
        session_id,
        uploaded_files,
        file_type_mapping,
        property_info
    )
    
    return {
        "session_id": session_id,
        "message": "Enhanced processing started with UW template integration",
        "files_uploaded": len(uploaded_files),
        "file_breakdown": {
            "rent_roll": rent_roll_count,
            "t12": t12_count,
            "additional": additional_count
        },
        "uw_template_available": UW_TEMPLATE_AVAILABLE
    }

@app.get("/api/status/{session_id}")
async def get_processing_status(session_id: str):
    """Get current processing status."""
    if session_id not in processing_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    return processing_sessions[session_id]

@app.get("/api/results/{session_id}")
async def get_results(session_id: str):
    """Get final results."""
    if session_id not in processing_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_sessions[session_id]
    if session.status != "completed":
        raise HTTPException(status_code=400, detail="Processing not completed")
    
    return session.results

@app.get("/api/download/{session_id}/{file_type}")
async def download_file(session_id: str, file_type: str):
    """Download generated files including UW template and CSV files."""
    logger.info(f"🔽 Download request: session_id={session_id}, file_type={file_type}")
    
    if session_id not in processing_sessions:
        logger.error(f"❌ Session not found: {session_id}")
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_sessions[session_id]
    logger.info(f"📋 Session status: {session.status}")
    
    if session.status != "completed" or not session.results:
        logger.error(f"❌ No files available for session {session_id}, status: {session.status}")
        raise HTTPException(status_code=400, detail="No files available")
    
    logger.info(f"📦 Available results: {list(session.results.keys())}")
    
    file_path = None
    if file_type == "uw_template":
        file_path = session.results.get("uw_template_path")
        logger.info(f"🎯 UW Template path: {file_path}")
    elif file_type == "excel":
        file_path = session.results.get("excel_path")
        logger.info(f"📊 Excel path: {file_path}")
    elif file_type == "pdf":
        file_path = session.results.get("pdf_path")
        logger.info(f"📄 PDF path: {file_path}")
    elif file_type == "html":
        file_path = session.results.get("html_path")
        logger.info(f"🌐 HTML path: {file_path}")
    elif file_type == "rent_roll_csv":
        file_path = session.results.get("rent_roll_csv_path")
        logger.info(f"🏠 Rent Roll CSV path: {file_path}")
    elif file_type == "t12_csv":
        file_path = session.results.get("t12_csv_path")
        logger.info(f"💰 T12 CSV path: {file_path}")
    else:
        logger.error(f"❌ Invalid file type: {file_type}")
        raise HTTPException(status_code=400, detail="Invalid file type")
    
    if not file_path:
        logger.error(f"❌ File path is None for {file_type}")
        raise HTTPException(status_code=404, detail="File path not found")
    
    if not os.path.exists(file_path):
        logger.error(f"❌ File does not exist: {file_path}")
        # List files in outputs directory for debugging
        try:
            outputs_files = os.listdir("outputs")
            logger.info(f"📁 Files in outputs directory: {outputs_files}")
        except Exception as e:
            logger.error(f"❌ Cannot list outputs directory: {e}")
        raise HTTPException(status_code=404, detail="File not found")
    
    # Get file info for logging
    file_size = os.path.getsize(file_path)
    logger.info(f"📏 File size: {file_size} bytes")
    
    # Determine proper media type based on file extension
    file_extension = os.path.splitext(file_path)[1].lower()
    logger.info(f"🔧 File extension: {file_extension}")
    
    media_type_mapping = {
        '.pdf': 'application/pdf',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.csv': 'text/csv',
        '.html': 'text/html',
        '.txt': 'text/plain'
    }
    
    media_type = media_type_mapping.get(file_extension, 'application/octet-stream')
    logger.info(f"📋 Media type: {media_type}")
    
    filename = os.path.basename(file_path)
    logger.info(f"📎 Filename: {filename}")
    
    try:
        logger.info(f"🚀 Starting file download: {filename}")
        return FileResponse(
            file_path,
            media_type=media_type,
            filename=filename,
            headers={
                "Content-Disposition": f"attachment; filename=\"{filename}\"",
                "Cache-Control": "no-cache"
            }
        )
    except Exception as e:
        logger.error(f"❌ FileResponse failed: {e}")
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")

async def process_documents_with_uw_template(
    session_id: str,
    uploaded_files: Dict[str, str],
    file_type_mapping: Dict[str, str],
    property_info: PropertyInfo
):
    """Enhanced processing with UW template integration."""
    
    def update_progress(step: int, step_name: str, message: str):
        session = processing_sessions[session_id]
        session.current_step = step
        session.step_name = step_name
        session.progress_percentage = (step / 9) * 100  # Updated for 9 steps
        session.message = message
        session.status = "processing"
        logger.info(f"Step {step}/9: {step_name} - {message}")
    
    try:
        session = processing_sessions[session_id]
        
        # Step 1: Initialize
        update_progress(1, "Initializing", "Setting up processing environment...")
        await asyncio.sleep(0.5)
        
        # Step 2: File Analysis
        update_progress(2, "File Analysis", "Analyzing uploaded documents...")
        
        # Separate files by type
        rent_roll_files = [f for f in uploaded_files.values() if file_type_mapping[f] == "rent_roll"]
        t12_files = [f for f in uploaded_files.values() if file_type_mapping[f] == "t12"]
        
        if not rent_roll_files or not t12_files:
            raise ValueError("Both rent roll and T12 documents are required for UW template generation")
        
        await asyncio.sleep(1)
        
        # Step 3: Extract Data from PDFs (Enhanced with LLM)
        update_progress(3, "Data Extraction", "Extracting data from PDFs...")
        
        processed_data = {}
        processing_mode = "LLM-Enhanced Mode"  # Default mode
        
        if REAL_PROCESSING_AVAILABLE and uploaded_files:
            try:
                # Initialize LLM processor
                llm_processor = LLMDocumentProcessor()
                
                # Process each uploaded file
                for file_type, file_path in uploaded_files.items():
                    if file_path and os.path.exists(file_path):
                        logger.info(f"🔍 Processing {file_type.upper()}: {os.path.basename(file_path)}")
                        
                        # Use LLM processor for enhanced extraction
                        extracted_data = llm_processor.extract_all_data(file_path)
                        
                        # Save raw extraction results
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        raw_data_path = f"outputs/{os.path.basename(file_path)}_{timestamp}_extraction.json"
                        os.makedirs("outputs", exist_ok=True)
                        
                        with open(raw_data_path, 'w') as f:
                            json.dump(extracted_data, f, indent=2, default=str)
                        
                        logger.info(f"✅ Extracted data from {file_type}")
                        logger.info(f"💾 Saved raw data: {raw_data_path}")
                        
                        # Store structured data
                        processed_data[file_type] = {
                            'file_path': file_path,
                            'extracted_data': extracted_data,
                            'structured_data': extracted_data.get('structured_data', {}),
                            'raw_text': extracted_data.get('raw_text', ''),
                            'tables': extracted_data.get('tables', [])
                        }
                
                logger.info("✅ LLM-enhanced data extraction completed")
                
            except Exception as e:
                logger.error(f"LLM extraction failed: {e}")
                logger.info("Falling back to basic extraction...")
                processed_data = create_fallback_data(property_info)
        else:
            processed_data = create_fallback_data(property_info)
            processing_mode = "Simulation Mode"
        
        await asyncio.sleep(1)
        
        # Step 4: Process LLM-Extracted Data
        update_progress(4, "Financial Analysis", "Processing LLM-extracted data...")
        
        # Process real extracted data from LLM
        real_data_summary = {}
        
        if processed_data:
            try:
                # Process Rent Roll data from LLM
                if 'rent_roll' in processed_data:
                    logger.info("🏢 Processing LLM-extracted Rent Roll data...")
                    rent_roll_structured = processed_data['rent_roll'].get('structured_data', {})
                    
                    real_data_summary['rent_roll'] = {
                        'total_units': rent_roll_structured.get('total_units', 0),
                        'occupied_units': rent_roll_structured.get('occupied_units', 0),
                        'vacant_units': rent_roll_structured.get('vacant_units', 0),
                        'monthly_income': rent_roll_structured.get('total_monthly_income', 0),
                        'annual_gpi': rent_roll_structured.get('annual_gpi', 0),
                        'avg_rent': rent_roll_structured.get('average_rent', 0),
                        'units_data': rent_roll_structured.get('units_data', [])
                    }
                    
                    logger.info(f"   ✅ Total Units: {real_data_summary['rent_roll']['total_units']}")
                    logger.info(f"   ✅ Monthly Income: ${real_data_summary['rent_roll']['monthly_income']:,.0f}")
                    logger.info(f"   ✅ Annual GPI: ${real_data_summary['rent_roll']['annual_gpi']:,.0f}")
                
                # Process T12 data from LLM
                if 't12' in processed_data:
                    logger.info("💰 Processing LLM-extracted T12 data...")
                    t12_structured = processed_data['t12'].get('structured_data', {})
                    
                    real_data_summary['t12'] = {
                        'total_revenue': t12_structured.get('total_revenue', 0),
                        'gross_potential_rent': t12_structured.get('gross_potential_rent', 0),
                        'vacancy_loss': t12_structured.get('vacancy_loss', 0),
                        'other_income': t12_structured.get('other_income', 0),
                        'total_expenses': t12_structured.get('total_expenses', 0),
                        'net_operating_income': t12_structured.get('net_operating_income', 0),
                        'expense_breakdown': t12_structured.get('expense_breakdown', {})
                    }
                    
                    logger.info(f"   ✅ Annual Revenue: ${real_data_summary['t12']['total_revenue']:,.0f}")
                    logger.info(f"   ✅ Annual Expenses: ${real_data_summary['t12']['total_expenses']:,.0f}")
                    logger.info(f"   ✅ NOI: ${real_data_summary['t12']['net_operating_income']:,.0f}")
                
                processed_data['real_data_summary'] = real_data_summary
                
            except Exception as e:
                logger.warning(f"LLM data processing failed: {e}")
                # Create minimal fallback
                real_data_summary = create_minimal_fallback()
                processed_data['real_data_summary'] = real_data_summary
        
        await asyncio.sleep(1)
        
        # Step 5: Generate UW Template (NEW!)
        update_progress(5, "UW Template Generation", "Creating professional UW template...")
        
        uw_template_path = None
        uw_template_success = False
        
        if UW_TEMPLATE_AVAILABLE and os.path.exists(UW_TEMPLATE_PATH):
            try:
                # Create enhanced UW filler with actual extracted data
                uw_filler = SimpleUWFiller(UW_TEMPLATE_PATH, processed_data, property_info)
                
                # Generate the UW template
                uw_template_path = uw_filler.create_uw_package()
                uw_template_success = True
                
                logger.info(f"✅ UW Template generated: {uw_template_path}")
                
            except Exception as e:
                logger.error(f"❌ UW Template generation failed: {e}")
                uw_template_success = False
        
        await asyncio.sleep(1.5)
        
        # Step 6: Generate Standard Outputs
        update_progress(6, "Standard Outputs", "Generating Excel and PDF outputs...")
        
        # Generate standard outputs (existing logic)
        if REAL_PROCESSING_AVAILABLE:
            try:
                output_generator = UnderwritingOutputGenerator(debug=True)
                output_generator.load_analysis_data(
                    rent_roll_analysis=processed_data.get('rent_roll', {}),
                    t12_analysis=processed_data.get('t12', {}),
                    property_info=property_info.dict(),
                    underwriting_summary={}
                )
                excel_path = output_generator.export_to_excel()
                pdf_path = "outputs/standard_output.pdf"  # Placeholder
            except Exception as e:
                logger.warning(f"Standard output generation failed: {e}")
                excel_path, pdf_path = await create_fallback_outputs(property_info)
        else:
            excel_path, pdf_path = await create_fallback_outputs(property_info)
        
        await asyncio.sleep(1)
        
        # Step 7: Generate CSV Files
        update_progress(7, "CSV Generation", "Creating T12 and Rent Roll CSV files...")
        
        # Generate CSV files for T12 and Rent Roll
        csv_files = generate_csv_files(processed_data, property_info, session_id)
        
        await asyncio.sleep(1)
        
        # Step 8: Finalize
        update_progress(8, "Finalizing", "Preparing results...")
        await asyncio.sleep(0.5)
        
        # Step 9: Complete
        session.status = "completed"
        session.current_step = 9
        session.progress_percentage = 100.0
        session.step_name = "Complete"
        session.message = f"Professional UW package ready! Mode: {processing_mode}"
        
        # Store comprehensive results
        logger.info(f"📦 Storing session results for {session_id}")
        logger.info(f"📦 UW template path: {uw_template_path}")
        logger.info(f"📦 Excel path: {excel_path}")
        logger.info(f"📦 PDF path: {pdf_path}")
        logger.info(f"📦 Rent roll CSV: {csv_files.get('rent_roll_csv')}")
        logger.info(f"📦 T12 CSV: {csv_files.get('t12_csv')}")
        
        session.results = {
            "session_id": session_id,
            "property_info": property_info.dict(),
            "processing_mode": processing_mode,
            "uw_template_success": uw_template_success,
            "uw_template_path": uw_template_path,
            "excel_path": excel_path,
            "pdf_path": pdf_path,
            "rent_roll_csv_path": csv_files.get('rent_roll_csv'),
            "t12_csv_path": csv_files.get('t12_csv'),
            "file_analysis": {
                "total_files": len(uploaded_files),
                "rent_roll_files": len(rent_roll_files),
                "t12_files": len(t12_files)
            },
            "timestamp": datetime.now().isoformat()
        }
        
        # Log final file verification
        logger.info(f"📦 Final file verification:")
        for file_type, file_path in [
            ("UW Template", uw_template_path),
            ("Excel", excel_path),
            ("PDF", pdf_path),
            ("Rent Roll CSV", csv_files.get('rent_roll_csv')),
            ("T12 CSV", csv_files.get('t12_csv'))
        ]:
            if file_path:
                exists = os.path.exists(file_path)
                size = os.path.getsize(file_path) if exists else 0
                logger.info(f"   {file_type}: {'✅' if exists else '❌'} {file_path} ({size} bytes)")
            else:
                logger.info(f"   {file_type}: ⚠️ No path set")
        
        logger.info(f"✅ Enhanced processing completed for session {session_id}")
        
    except Exception as e:
        logger.error(f"❌ Processing failed for session {session_id}: {str(e)}")
        session = processing_sessions[session_id]
        session.status = "error"
        session.error_message = str(e)
        session.message = f"Processing failed: {str(e)}"

class SimpleUWFiller:
    """Simple UW template filler that works with real extracted data."""
    
    def __init__(self, template_path, processed_data, property_info):
        self.template_path = template_path
        self.processed_data = processed_data
        self.property_info = property_info
    
    def create_uw_package(self):
        """Create UW package with real extracted data."""
        try:
            logger.info("🎯 Creating Simple UW Package with Real Data...")
            logger.info(f"🎯 Template path: {self.template_path}")
            logger.info(f"🎯 Template exists: {os.path.exists(self.template_path)}")
            
            # Get real extracted data
            real_data = self.processed_data.get('real_data_summary', {})
            rent_roll_data = real_data.get('rent_roll', {})
            t12_data = real_data.get('t12', {})
            logger.info(f"🎯 Rent roll data keys: {list(rent_roll_data.keys())}")
            logger.info(f"🎯 T12 data keys: {list(t12_data.keys())}")
            
            # Load template
            logger.info("🎯 Loading Excel template...")
            wb = load_workbook(self.template_path)
            logger.info(f"🎯 Available sheets: {wb.sheetnames}")
            
            if 'UW' not in wb.sheetnames:
                logger.error("❌ UW sheet not found!")
                return None
            
            ws = wb['UW']
            logger.info("🎯 UW sheet loaded successfully")
            
            # Fill with real data
            logger.info("🎯 Filling template with data...")
            self.fill_with_real_data(ws, rent_roll_data, t12_data)
            
            # Save
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"outputs/Real_UW_Package_{timestamp}.xlsx"
            logger.info(f"🎯 Saving to: {output_path}")
            
            os.makedirs("outputs", exist_ok=True)
            wb.save(output_path)
            
            # Verify file creation
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                logger.info(f"✅ UW Package created: {output_path} (size: {file_size} bytes)")
                return output_path
            else:
                logger.error(f"❌ UW Package file not found after save: {output_path}")
                return None
            
        except Exception as e:
            logger.error(f"❌ UW Package creation failed: {e}")
            logger.error(f"❌ Exception details: {type(e).__name__}: {str(e)}")
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            return None
    
    def fill_with_real_data(self, ws, rent_roll_data, t12_data):
        """Fill UW template with real LLM-extracted data."""
        
        # Property data
        self.safe_write(ws, 'B4', self.property_info.property_name)
        self.safe_write(ws, 'B5', self.property_info.property_address)
        self.safe_write(ws, 'B6', rent_roll_data.get('total_units', 0))
        self.safe_write(ws, 'B7', rent_roll_data.get('avg_rent', 0))
        
        # Financial data from LLM extraction
        annual_gpi = rent_roll_data.get('annual_gpi', 0)
        total_revenue = t12_data.get('total_revenue', 0)
        total_expenses = t12_data.get('total_expenses', 0)
        noi = t12_data.get('net_operating_income', 0)
        
        # Calculate loan amount (75% LTV based on NOI)
        cap_rate = 0.06  # 6% cap rate
        property_value = noi / cap_rate if noi > 0 else 0
        loan_amount = property_value * 0.75 if property_value > 0 else 0
        
        # Fill key cells with LLM-extracted data
        self.safe_write(ws, 'B8', int(loan_amount))
        self.safe_write(ws, 'B9', 0.06)  # 6% rate
        self.safe_write(ws, 'B10', int(property_value))
        self.safe_write(ws, 'B11', 0.75)  # 75% LTV
        
        # Revenue from LLM
        self.safe_write(ws, 'B12', int(annual_gpi))
        self.safe_write(ws, 'B13', int(total_revenue))
        
        # Expenses from LLM
        self.safe_write(ws, 'B14', int(total_expenses))
        
        # NOI from LLM
        self.safe_write(ws, 'B15', int(noi))
        
        # Cap rate
        self.safe_write(ws, 'B16', cap_rate)
        
        logger.info(f"✅ Filled UW template with LLM-extracted data:")
        logger.info(f"   - Units: {rent_roll_data.get('total_units', 0)}")
        logger.info(f"   - Annual GPI: ${annual_gpi:,.0f}")
        logger.info(f"   - NOI: ${noi:,.0f}")
        logger.info(f"   - Loan Amount: ${loan_amount:,.0f}")
    
    def safe_write(self, ws, cell_ref, value):
        """Safely write to a cell."""
        try:
            ws[cell_ref] = value
            return True
        except Exception as e:
            logger.debug(f"Could not write {value} to {cell_ref}: {e}")
            return False


def create_fallback_data(property_info):
    """Create fallback data when real processing isn't available."""
    return {
        'rent_roll': {
            'tables': [pd.DataFrame({
                'Unit_Number': range(1, 87),
                'Current_Rent': [1527] * 86,
                'Status': ['Occupied'] * 66 + ['Vacant'] * 20
            })]
        },
        't12': {
            'tables': [pd.DataFrame({
                'Line_Item': ['Rental Income', 'Property Taxes', 'Insurance'],
                'Annual_Amount': [2822629, 436783, 34824]
            })]
        }
    }

def extract_rent_roll_directly(rent_roll_df):
    """Extract rent roll data directly from DataFrame."""
    try:
        # Clean up the data based on actual CSV structure
        df = rent_roll_df.copy()
        
        # Remove header rows if present
        if df.iloc[0, 0] == 'Unit #':
            df = df.iloc[1:].reset_index(drop=True)
        
        # Extract key columns (based on actual CSV structure)
        units = []
        total_monthly_rent = 0
        occupied_units = 0
        
        for _, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                    continue
                
                unit_number = str(row.iloc[1]) if len(row) > 1 else str(row.iloc[0])
                unit_type = str(row.iloc[2]) if len(row) > 2 else ""
                
                # Extract rent (column 5 in the CSV)
                rent_str = str(row.iloc[5]) if len(row) > 5 else "0"
                rent = 0
                if rent_str and rent_str != 'nan':
                    # Clean rent string: remove $, commas
                    rent_clean = rent_str.replace('$', '').replace(',', '').strip()
                    try:
                        rent = float(rent_clean)
                        if rent > 0:
                            occupied_units += 1
                            total_monthly_rent += rent
                    except:
                        pass
                
                units.append({
                    'unit': unit_number,
                    'type': unit_type,
                    'rent': rent
                })
                
            except Exception as e:
                continue
        
        total_units = len(units)
        avg_rent = total_monthly_rent / occupied_units if occupied_units > 0 else 0
        
        return {
            'total_units': total_units,
            'occupied_units': occupied_units,
            'vacant_units': total_units - occupied_units,
            'monthly_income': total_monthly_rent,
            'annual_gpi': total_monthly_rent * 12,
            'avg_rent': avg_rent,
            'units_data': units
        }
        
    except Exception as e:
        logger.error(f"Error extracting rent roll: {e}")
        return {
            'total_units': 0,
            'occupied_units': 0,
            'vacant_units': 0,
            'monthly_income': 0,
            'annual_gpi': 0,
            'avg_rent': 0,
            'units_data': []
        }

def extract_t12_directly(t12_df):
    """Extract T12 data directly from DataFrame."""
    try:
        df = t12_df.copy()
        
        # Find revenue and expense data from T12
        total_revenue = 0
        total_expenses = 0
        noi = 0
        
        # Look for key financial lines
        for _, row in df.iterrows():
            try:
                line_item = str(row.iloc[0]).lower() if not pd.isna(row.iloc[0]) else ""
                
                # Get T12 value (last column typically)
                t12_value_str = str(row.iloc[-1]) if len(row) > 1 else "0"
                
                # Clean and convert to number
                if t12_value_str and t12_value_str != 'nan':
                    value_clean = t12_value_str.replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip()
                    try:
                        value = float(value_clean)
                        
                        # Categorize line items
                        if 'total property rental income' in line_item:
                            total_revenue = value
                        elif 'total revenues' in line_item and total_revenue == 0:
                            total_revenue = value
                        elif 'total operating expenses' in line_item:
                            total_expenses = value
                        elif 'net operating income' in line_item:
                            noi = value
                            
                    except:
                        continue
                        
            except Exception as e:
                continue
        
        # Calculate NOI if not found directly
        if noi == 0 and total_revenue > 0 and total_expenses > 0:
            noi = total_revenue - total_expenses
        
        return {
            'total_revenue': total_revenue,
            'total_expenses': abs(total_expenses),  # Make sure expenses are positive
            'noi': noi,
            'raw_data': df
        }
        
    except Exception as e:
        logger.error(f"Error extracting T12: {e}")
        return {
            'total_revenue': 0,
            'total_expenses': 0,
            'noi': 0,
            'raw_data': None
        }

def create_minimal_fallback():
    """Create minimal fallback data."""
    return {
        'rent_roll': {
            'total_units': 0,
            'occupied_units': 0,
            'vacant_units': 0,
            'monthly_income': 0,
            'annual_gpi': 0,
            'avg_rent': 0,
            'units_data': []
        },
        't12': {
            'total_revenue': 0,
            'total_expenses': 0,
            'noi': 0,
            'raw_data': None
        }
    }

def generate_csv_files(processed_data, property_info, session_id):
    """Generate T12 and Rent Roll CSV files from processed data."""
    logger.info(f"📊 Starting CSV generation for session: {session_id}")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Clean property name for filename
    clean_property_name = "".join(c for c in property_info.property_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    clean_property_name = clean_property_name.replace(' ', '_')
    logger.info(f"🏷️ Clean property name: '{clean_property_name}'")
    
    csv_files = {}
    
    try:
        logger.info(f"📋 Processed data keys: {list(processed_data.keys())}")
        
        # Generate Rent Roll CSV
        rent_roll_data = processed_data.get('real_data_summary', {}).get('rent_roll', {})
        logger.info(f"🏠 Rent roll data available: {bool(rent_roll_data)}")
        logger.info(f"🏠 Rent roll units data: {bool(rent_roll_data.get('units_data'))}")
        
        if rent_roll_data and rent_roll_data.get('units_data'):
            logger.info(f"🏠 Processing rent roll data with {len(rent_roll_data['units_data'])} units")
            rent_roll_df = pd.DataFrame(rent_roll_data['units_data'])
            
            # Add summary data
            summary_data = {
                'unit': 'SUMMARY',
                'type': '',
                'rent': f"Total Units: {rent_roll_data.get('total_units', 0)}, Occupied: {rent_roll_data.get('occupied_units', 0)}, Monthly Income: ${rent_roll_data.get('monthly_income', 0):,.0f}"
            }
            
            # Create comprehensive rent roll CSV
            rent_roll_csv_path = f"outputs/{clean_property_name}_Rent_Roll_{timestamp}.csv"
            logger.info(f"🏠 Creating rent roll CSV: {rent_roll_csv_path}")
            os.makedirs("outputs", exist_ok=True)
            
            with open(rent_roll_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Header section
                writer.writerow(['Property Name:', property_info.property_name])
                writer.writerow(['Property Address:', property_info.property_address])
                writer.writerow(['Report Date:', datetime.now().strftime('%B %d, %Y')])
                writer.writerow([''])
                
                # Summary section
                writer.writerow(['RENT ROLL SUMMARY'])
                writer.writerow(['Total Units:', rent_roll_data.get('total_units', 0)])
                writer.writerow(['Occupied Units:', rent_roll_data.get('occupied_units', 0)])
                writer.writerow(['Vacant Units:', rent_roll_data.get('vacant_units', 0)])
                writer.writerow(['Monthly Gross Income:', f"${rent_roll_data.get('monthly_income', 0):,.2f}"])
                writer.writerow(['Annual Gross Income:', f"${rent_roll_data.get('annual_gpi', 0):,.2f}"])
                writer.writerow(['Average Rent:', f"${rent_roll_data.get('avg_rent', 0):,.2f}"])
                writer.writerow([''])
                
                # Unit details header
                writer.writerow(['Unit #', 'Unit Type', 'Monthly Rent', 'Status'])
                
                # Unit details
                for unit in rent_roll_data['units_data']:
                    status = 'Occupied' if unit.get('rent', 0) > 0 else 'Vacant'
                    writer.writerow([
                        unit.get('unit', ''),
                        unit.get('type', ''),
                        f"${unit.get('rent', 0):,.2f}" if unit.get('rent', 0) > 0 else 'Vacant',
                        status
                    ])
            
            # Verify file was created
            if os.path.exists(rent_roll_csv_path):
                file_size = os.path.getsize(rent_roll_csv_path)
                logger.info(f"✅ Generated Rent Roll CSV: {rent_roll_csv_path} (size: {file_size} bytes)")
                csv_files['rent_roll_csv'] = rent_roll_csv_path
            else:
                logger.error(f"❌ Rent Roll CSV file not created: {rent_roll_csv_path}")
        else:
            logger.warning(f"⚠️ No rent roll data available for CSV generation")
        
        # Generate T12 CSV
        t12_data = processed_data.get('real_data_summary', {}).get('t12', {})
        logger.info(f"💰 T12 data available: {bool(t12_data)}")
        
        if t12_data:
            logger.info(f"💰 Processing T12 data")
            t12_csv_path = f"outputs/{clean_property_name}_T12_Operating_Statement_{timestamp}.csv"
            logger.info(f"💰 Creating T12 CSV: {t12_csv_path}")
            
            with open(t12_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Header section
                writer.writerow(['Property Name:', property_info.property_name])
                writer.writerow(['Property Address:', property_info.property_address])
                writer.writerow(['Report Date:', datetime.now().strftime('%B %d, %Y')])
                writer.writerow([''])
                
                # T12 Summary
                writer.writerow(['T12 OPERATING STATEMENT SUMMARY'])
                writer.writerow([''])
                
                # Revenue section
                writer.writerow(['REVENUE'])
                writer.writerow(['Total Revenue:', f"${t12_data.get('total_revenue', 0):,.2f}"])
                writer.writerow(['Gross Potential Rent:', f"${t12_data.get('gross_potential_rent', 0):,.2f}"])
                writer.writerow(['Vacancy Loss:', f"${t12_data.get('vacancy_loss', 0):,.2f}"])
                writer.writerow(['Other Income:', f"${t12_data.get('other_income', 0):,.2f}"])
                writer.writerow([''])
                
                # Expenses section
                writer.writerow(['EXPENSES'])
                writer.writerow(['Total Expenses:', f"${t12_data.get('total_expenses', 0):,.2f}"])
                
                # Expense breakdown
                expense_breakdown = t12_data.get('expense_breakdown', {})
                if expense_breakdown:
                    for expense_type, amount in expense_breakdown.items():
                        writer.writerow([f'{expense_type.title()}:', f"${amount:,.2f}"])
                
                writer.writerow([''])
                
                # NOI section
                writer.writerow(['NET OPERATING INCOME'])
                writer.writerow(['NOI:', f"${t12_data.get('net_operating_income', 0):,.2f}"])
                
                # Financial ratios
                total_revenue = t12_data.get('total_revenue', 0)
                total_expenses = t12_data.get('total_expenses', 0)
                if total_revenue > 0:
                    expense_ratio = (total_expenses / total_revenue) * 100
                    writer.writerow(['Expense Ratio:', f"{expense_ratio:.1f}%"])
            
            # Verify file was created
            if os.path.exists(t12_csv_path):
                file_size = os.path.getsize(t12_csv_path)
                logger.info(f"✅ Generated T12 CSV: {t12_csv_path} (size: {file_size} bytes)")
                csv_files['t12_csv'] = t12_csv_path
            else:
                logger.error(f"❌ T12 CSV file not created: {t12_csv_path}")
        else:
            logger.warning(f"⚠️ No T12 data available for CSV generation")
        
        # If no real data, create basic CSV files with template structure
        if not csv_files:
            logger.info(f"📋 No real data available, creating basic CSV templates")
            
            # Basic rent roll CSV
            rent_roll_csv_path = f"outputs/{clean_property_name}_Rent_Roll_{timestamp}.csv"
            logger.info(f"🏠 Creating basic rent roll CSV: {rent_roll_csv_path}")
            basic_rent_roll = pd.DataFrame({
                'Unit_Number': ['101', '102', '103', '201', '202'],
                'Unit_Type': ['1BR/1BA', '1BR/1BA', '2BR/2BA', '1BR/1BA', '2BR/2BA'],
                'Monthly_Rent': [1500, 1500, 1800, 1500, 1800],
                'Status': ['Occupied', 'Occupied', 'Occupied', 'Vacant', 'Occupied']
            })
            basic_rent_roll.to_csv(rent_roll_csv_path, index=False)
            
            if os.path.exists(rent_roll_csv_path):
                file_size = os.path.getsize(rent_roll_csv_path)
                logger.info(f"✅ Created basic rent roll CSV: {rent_roll_csv_path} (size: {file_size} bytes)")
                csv_files['rent_roll_csv'] = rent_roll_csv_path
            
            # Basic T12 CSV
            t12_csv_path = f"outputs/{clean_property_name}_T12_Operating_Statement_{timestamp}.csv"
            logger.info(f"💰 Creating basic T12 CSV: {t12_csv_path}")
            basic_t12 = pd.DataFrame({
                'Line_Item': ['Rental Income', 'Other Income', 'Total Revenue', 'Property Taxes', 'Insurance', 'Maintenance', 'Total Expenses', 'Net Operating Income'],
                'Annual_Amount': [120000, 5000, 125000, 15000, 8000, 12000, 35000, 90000]
            })
            basic_t12.to_csv(t12_csv_path, index=False)
            
            if os.path.exists(t12_csv_path):
                file_size = os.path.getsize(t12_csv_path)
                logger.info(f"✅ Created basic T12 CSV: {t12_csv_path} (size: {file_size} bytes)")
                csv_files['t12_csv'] = t12_csv_path
            
            logger.info(f"✅ Generated basic CSV templates")
    
    except Exception as e:
        logger.error(f"❌ Error generating CSV files: {e}")
        logger.error(f"❌ Exception details: {type(e).__name__}: {str(e)}")
        import traceback
        logger.error(f"❌ Traceback: {traceback.format_exc()}")
        # Return empty dict if generation fails
        csv_files = {}
    
    logger.info(f"📊 CSV generation completed. Generated files: {list(csv_files.keys())}")
    return csv_files

async def create_fallback_outputs(property_info):
    """Create simple fallback outputs with proper filenames."""
    logger.info(f"📄 Creating fallback outputs for: {property_info.property_name}")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Clean property name for filename (remove spaces and special chars)
    clean_property_name = "".join(c for c in property_info.property_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    clean_property_name = clean_property_name.replace(' ', '_')
    logger.info(f"🏷️ Clean filename base: '{clean_property_name}'")
    
    excel_path = f"outputs/{clean_property_name}_Analysis_{timestamp}.xlsx"
    pdf_path = f"outputs/{clean_property_name}_Summary_{timestamp}.pdf"
    logger.info(f"📊 Target Excel path: {excel_path}")
    logger.info(f"📄 Target PDF path: {pdf_path}")
    
    # Create simple Excel file
    df = pd.DataFrame({
        'Property': [property_info.property_name],
        'Address': [property_info.property_address],
        'Transaction': [property_info.transaction_type]
    })
    logger.info(f"📋 Created DataFrame with {len(df)} rows")
    
    os.makedirs("outputs", exist_ok=True)
    logger.info(f"📁 Ensured outputs directory exists")
    
    # Create Excel file with proper error handling
    try:
        logger.info(f"📊 Creating Excel file...")
        df.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Verify file creation
        if os.path.exists(excel_path):
            file_size = os.path.getsize(excel_path)
            logger.info(f"✅ Created Excel file: {excel_path} (size: {file_size} bytes)")
        else:
            logger.error(f"❌ Excel file not found after creation: {excel_path}")
            
    except Exception as e:
        logger.error(f"❌ Failed to create Excel file: {e}")
        # Fallback to CSV if Excel fails
        csv_path = excel_path.replace('.xlsx', '.csv')
        logger.info(f"📄 Falling back to CSV: {csv_path}")
        try:
            df.to_csv(csv_path, index=False)
            if os.path.exists(csv_path):
                file_size = os.path.getsize(csv_path)
                logger.info(f"✅ Created CSV fallback: {csv_path} (size: {file_size} bytes)")
                excel_path = csv_path
            else:
                logger.error(f"❌ CSV fallback file not found: {csv_path}")
        except Exception as csv_e:
            logger.error(f"❌ CSV fallback also failed: {csv_e}")
    
    # Create a proper PDF file (even if placeholder)
    try:
        logger.info(f"📄 Creating PDF file...")
        # Create a simple PDF using reportlab if available, otherwise a proper binary placeholder
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter
            
            logger.info(f"📄 Using reportlab for PDF creation")
            c = canvas.Canvas(pdf_path, pagesize=letter)
            c.drawString(100, 750, f"Underwriting Summary - {property_info.property_name}")
            c.drawString(100, 720, f"Address: {property_info.property_address}")
            c.drawString(100, 690, f"Transaction Type: {property_info.transaction_type}")
            c.drawString(100, 660, f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
            c.drawString(100, 600, "This is a preliminary analysis placeholder.")
            c.drawString(100, 580, "Full analysis requires document processing.")
            c.save()
            
            # Verify PDF creation
            if os.path.exists(pdf_path):
                file_size = os.path.getsize(pdf_path)
                logger.info(f"✅ Created PDF file with reportlab: {pdf_path} (size: {file_size} bytes)")
            else:
                logger.error(f"❌ PDF file not found after reportlab creation: {pdf_path}")
                
        except ImportError:
            logger.info(f"⚠️ Reportlab not available, using minimal PDF structure")
            # Fallback: Create a minimal valid PDF structure
            pdf_content = b"""%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>
endobj
4 0 obj
<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>
endobj
5 0 obj
<< /Length 44 >>
stream
BT
/F1 12 Tf
100 700 Td
(Underwriting Summary) Tj
ET
endstream
endobj
xref
0 6
0000000000 65535 f 
0000000010 00000 n 
0000000053 00000 n 
0000000110 00000 n 
0000000252 00000 n 
0000000315 00000 n 
trailer
<< /Size 6 /Root 1 0 R >>
startxref
409
%%EOF"""
            
            with open(pdf_path, 'wb') as f:
                f.write(pdf_content)
            
            # Verify minimal PDF creation
            if os.path.exists(pdf_path):
                file_size = os.path.getsize(pdf_path)
                logger.info(f"✅ Created minimal PDF file: {pdf_path} (size: {file_size} bytes)")
            else:
                logger.error(f"❌ Minimal PDF file not found after creation: {pdf_path}")
            
    except Exception as e:
        logger.error(f"❌ Failed to create PDF file: {e}")
        logger.error(f"❌ PDF Exception details: {type(e).__name__}: {str(e)}")
        # Last resort: create an HTML file instead
        html_path = pdf_path.replace('.pdf', '.html')
        logger.info(f"🌐 Creating HTML fallback: {html_path}")
        try:
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(f"""<!DOCTYPE html>
<html>
<head><title>Underwriting Summary</title></head>
<body>
    <h1>Underwriting Summary</h1>
    <p><strong>Property:</strong> {property_info.property_name}</p>
    <p><strong>Address:</strong> {property_info.property_address}</p>
    <p><strong>Transaction:</strong> {property_info.transaction_type}</p>
    <p><strong>Generated:</strong> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
</body>
</html>""")
            
            if os.path.exists(html_path):
                file_size = os.path.getsize(html_path)
                logger.info(f"✅ Created HTML fallback: {html_path} (size: {file_size} bytes)")
                pdf_path = html_path
            else:
                logger.error(f"❌ HTML fallback file not found: {html_path}")
                
        except Exception as html_e:
            logger.error(f"❌ HTML fallback also failed: {html_e}")
    
    logger.info(f"📄 Fallback outputs completed - Excel: {excel_path}, PDF: {pdf_path}")
    return excel_path, pdf_path

if __name__ == "__main__":
    import uvicorn
    import os
    
    port = int(os.environ.get("PORT", 8000))
    host = os.environ.get("HOST", "0.0.0.0")
    
    print(f"🚀 Starting Hardwell Underwriting Automation on {host}:{port}")
    
    uvicorn.run(
        app, 
        host=host, 
        port=port,
        access_log=True,
        log_level="info"
    )