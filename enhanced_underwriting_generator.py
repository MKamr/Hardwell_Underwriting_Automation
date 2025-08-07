#!/usr/bin/env python3
"""
Enhanced Professional Underwriting Generator
Creates comprehensive underwriting analysis matching industry standards
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from document_processor import DocumentProcessor
from underwriting_analyzer import UnderwritingAnalyzer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import logging

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class EnhancedUnderwritingGenerator:
    """Enhanced generator for professional underwriting packages."""
    
    def __init__(self, debug=False):
        self.debug = debug
        self.processor = DocumentProcessor(debug=debug)
        self.analyzer = UnderwritingAnalyzer(debug=debug)
        
        # Rulebook configuration
        self.config = {
            'vacancy_rates': {
                'class_a': 0.05,  # 5%
                'class_b': 0.07,  # 7%
                'class_c': 0.10   # 10%
            },
            'rm_minimums': {
                (0, 10): 500,    # $500/unit for 0-10 years
                (10, 20): 600,   # $600/unit for 10-20 years
                (20, 30): 700,   # $700/unit for 20-30 years
                (30, 50): 800,   # $800/unit for 30-50 years
                (50, 100): 1000  # $1000/unit for 50+ years
            },
            'management_fee_tiers': [
                (0, 500000, 0.05),      # 5% for GPI 0-500k
                (500000, 750000, 0.045), # 4.5% for GPI 500k-750k
                (750000, 1000000, 0.04), # 4% for GPI 750k-1M
                (1000000, 1500000, 0.035), # 3.5% for GPI 1M-1.5M
                (1500000, float('inf'), 0.03) # 3% for GPI 1.5M+
            ],
            'expense_adjustments': {
                'property_taxes_refinance': 1.075,  # +7.5% for refinance
                'property_taxes_purchase': 1.0,     # No adjustment for purchase
                'insurance': 1.05,                  # +5%
                'utilities': 1.02,                  # +2%
                'minimum_expense_ratio': 0.28       # 28% minimum
            },
            'replacement_reserves': 250  # $250/unit
        }
    
    def generate_professional_package(self, rent_roll_path, t12_path, property_info):
        """Generate comprehensive professional underwriting package."""
        
        print("🚀 Generating Professional Underwriting Package...")
        print("=" * 80)
        
        # Extract and process data
        print("📊 Processing Rent Roll...")
        rent_roll_data = self._process_rent_roll(rent_roll_path)
        
        print("💰 Processing T12...")
        t12_data = self._process_t12(t12_path)
        
        print("📈 Generating Comprehensive Analysis...")
        analysis = self._generate_comprehensive_analysis(rent_roll_data, t12_data, property_info)
        
        print("📊 Creating Professional Excel Package...")
        excel_path = self._create_professional_excel(analysis, rent_roll_data)
        
        print("📄 Creating Professional PDF Package...")
        pdf_path = self._create_professional_pdf(analysis, excel_path)
        
        return {
            'excel_path': excel_path,
            'pdf_path': pdf_path,
            'analysis': analysis,
            'rent_roll_data': rent_roll_data,
            't12_data': t12_data
        }
    
    def _process_rent_roll(self, rent_roll_path):
        """Process rent roll with enhanced data extraction."""
        
        results = self.processor.process_document(rent_roll_path)
        rent_roll_df = results['tables'][0] if results.get('tables') else None
        
        if rent_roll_df is None:
            raise ValueError("Failed to extract rent roll data")
        
        # Clean and structure data
        cleaned_data = []
        for idx, row in rent_roll_df.iterrows():
            if idx == 0:  # Skip header
                continue
                
            unit_number = str(row.iloc[0]).strip()
            if not unit_number or unit_number == 'nan' or 'Unit' in unit_number:
                continue
            
            # Extract rent
            rent_str = str(row.iloc[5]).replace('$', '').replace(',', '').strip()
            try:
                current_rent = float(rent_str) if rent_str and rent_str != 'nan' else 0
            except:
                current_rent = 0
            
            # Extract square footage
            sqft_str = str(row.iloc[3]).replace(',', '').strip()
            try:
                square_feet = float(sqft_str) if sqft_str and sqft_str != 'nan' else 1187
            except:
                square_feet = 1187
            
            # Determine status
            tenant_name = str(row.iloc[4]).strip()
            status = 'Occupied' if tenant_name and tenant_name != 'nan' else 'Vacant'
            
            cleaned_data.append({
                'Unit_Number': unit_number,
                'Unit_Type': str(row.iloc[2]).strip(),
                'Square_Feet': square_feet,
                'Current_Rent': current_rent,
                'Status': status,
                'Tenant_Name': tenant_name if status == 'Occupied' else '',
                'Lease_End_Date': str(row.iloc[10]).strip() if len(row) > 10 else ''
            })
        
        rent_roll_df = pd.DataFrame(cleaned_data)
        print(f"   ✅ Extracted {len(rent_roll_df)} units")
        
        return rent_roll_df
    
    def _process_t12(self, t12_path):
        """Process T12 with enhanced financial extraction."""
        
        results = self.processor.process_document(t12_path)
        t12_df = results['tables'][0] if results.get('tables') else None
        
        if t12_df is None:
            raise ValueError("Failed to extract T12 data")
        
        # Extract financial metrics
        financial_data = {}
        
        for idx, row in t12_df.iterrows():
            row_text = str(row.iloc[0]).strip()
            amount_col = row.iloc[-1] if len(row) > 1 else None
            
            # Map T12 line items to standardized names
            if 'Gross Potential Rent' in row_text:
                financial_data['gross_potential_rents'] = self._extract_amount(amount_col)
            elif 'Loss to Lease' in row_text:
                financial_data['loss_to_lease'] = self._extract_amount(amount_col)
            elif 'Vacancy' in row_text and 'Loss' in row_text:
                financial_data['vacancy_loss'] = self._extract_amount(amount_col)
            elif 'TOTAL PROPERTY RENTAL INCOME' in row_text:
                financial_data['rental_income'] = self._extract_amount(amount_col)
            elif 'TOTAL OTHER INCOME' in row_text:
                financial_data['other_income'] = self._extract_amount(amount_col)
            elif 'Property Taxes' in row_text:
                financial_data['property_taxes'] = self._extract_amount(amount_col)
            elif 'Insurance' in row_text and 'Premium' in row_text:
                financial_data['insurance'] = self._extract_amount(amount_col)
            elif 'TOTAL UTILITIES' in row_text:
                financial_data['utilities'] = self._extract_amount(amount_col)
            elif 'Maintenance' in row_text and 'Repair' in row_text:
                financial_data['maintenance_repairs'] = self._extract_amount(amount_col)
            elif 'Management Fee' in row_text:
                financial_data['management_fees'] = self._extract_amount(amount_col)
            elif 'TOTAL OPERATING EXPENSES' in row_text:
                financial_data['total_operating_expenses'] = self._extract_amount(amount_col)
            elif 'NET OPERATING INCOME' in row_text:
                financial_data['net_operating_income'] = self._extract_amount(amount_col)
        
        print(f"   ✅ Extracted financial metrics")
        return financial_data
    
    def _extract_amount(self, amount_str):
        """Extract numeric amount from string."""
        if pd.isna(amount_str):
            return 0
        
        amount_str = str(amount_str).replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip()
        try:
            return float(amount_str) if amount_str and amount_str != 'nan' else 0
        except:
            return 0
    
    def _generate_comprehensive_analysis(self, rent_roll_df, t12_data, property_info):
        """Generate comprehensive underwriting analysis with rulebook compliance."""
        
        # Property characteristics
        total_units = len(rent_roll_df)
        occupied_units = len(rent_roll_df[rent_roll_df['Status'] == 'Occupied'])
        vacant_units = total_units - occupied_units
        average_rent = rent_roll_df['Current_Rent'].mean()
        average_sqft = rent_roll_df['Square_Feet'].mean()
        rent_per_sqft = average_rent / average_sqft if average_sqft > 0 else 0
        
        # Unit type analysis
        unit_types = rent_roll_df['Unit_Type'].value_counts().to_dict()
        
        # Income analysis
        gross_potential_rents = t12_data.get('gross_potential_rents', 0)
        rental_income = t12_data.get('rental_income', 0)
        other_income = t12_data.get('other_income', 0)
        
        # Calculate vacancy rate
        actual_vacancy_rate = (gross_potential_rents - rental_income) / gross_potential_rents if gross_potential_rents > 0 else 0
        
        # Apply rulebook vacancy rate (minimum 5%)
        underwriting_vacancy_rate = max(0.05, actual_vacancy_rate)
        
        # Calculate underwritten income
        uw_gross_potential_income = gross_potential_rents
        uw_vacancy_loss = uw_gross_potential_income * underwriting_vacancy_rate
        uw_effective_gross_income = uw_gross_potential_income - uw_vacancy_loss + other_income
        
        # Expense analysis with rulebook adjustments
        property_taxes = t12_data.get('property_taxes', 0)
        insurance = t12_data.get('insurance', 0)
        utilities = t12_data.get('utilities', 0)
        maintenance_repairs = t12_data.get('maintenance_repairs', 0)
        management_fees = t12_data.get('management_fees', 0)
        
        # Apply rulebook adjustments
        transaction_type = getattr(property_info, 'transaction_type', 'refinance')
        if transaction_type == 'refinance':
            uw_property_taxes = property_taxes * self.config['expense_adjustments']['property_taxes_refinance']
        else:
            uw_property_taxes = property_taxes * self.config['expense_adjustments']['property_taxes_purchase']
        
        uw_insurance = insurance * self.config['expense_adjustments']['insurance']
        uw_utilities = utilities * self.config['expense_adjustments']['utilities']
        
        # R&M minimum based on property age
        property_age = getattr(property_info, 'property_age', 25)  # Default 25 years
        rm_minimum = self._calculate_rm_minimum(total_units, property_age)
        uw_maintenance_repairs = max(maintenance_repairs, rm_minimum)
        
        # Management fee tiers
        uw_management_fees = self._calculate_management_fees(uw_gross_potential_income)
        
        # Replacement reserves
        uw_replacement_reserves = total_units * self.config['replacement_reserves']
        
        # Total expenses
        uw_total_expenses = (uw_property_taxes + uw_insurance + uw_utilities + 
                           uw_maintenance_repairs + uw_management_fees + uw_replacement_reserves)
        
        # Ensure minimum expense ratio
        min_expenses = uw_effective_gross_income * self.config['expense_adjustments']['minimum_expense_ratio']
        uw_total_expenses = max(uw_total_expenses, min_expenses)
        
        # Net Operating Income
        uw_net_operating_income = uw_effective_gross_income - uw_total_expenses
        expense_ratio = uw_total_expenses / uw_effective_gross_income if uw_effective_gross_income > 0 else 0
        
        # Valuation analysis
        market_cap_rate = 0.065  # 6.5% default
        estimated_value = uw_net_operating_income / market_cap_rate if market_cap_rate > 0 else 0
        price_per_unit = estimated_value / total_units if total_units > 0 else 0
        
        # Rent analysis
        rent_analysis = self._generate_rent_analysis(rent_roll_df)
        
        return {
            'property_characteristics': {
                'property_name': getattr(property_info, 'property_name', 'Property'),
                'property_address': getattr(property_info, 'property_address', 'Address'),
                'total_units': total_units,
                'occupied_units': occupied_units,
                'vacant_units': vacant_units,
                'unit_types': unit_types,
                'average_rent': average_rent,
                'average_sqft': average_sqft,
                'rent_per_sqft': rent_per_sqft,
                'property_age': property_age,
                'transaction_type': transaction_type
            },
            'loan_terms': {
                'loan_amount': getattr(property_info, 'loan_amount', estimated_value * 0.75),
                'loan_to_value': 0.75,
                'interest_rate': 0.055,  # 5.5%
                'amortization': 30,
                'term': 10
            },
            'income_analysis': {
                'gross_potential_income': uw_gross_potential_income,
                'vacancy_loss': uw_vacancy_loss,
                'vacancy_rate': underwriting_vacancy_rate,
                'effective_gross_income': uw_effective_gross_income,
                'rental_income': rental_income,
                'other_income': other_income
            },
            'expense_analysis': {
                'property_taxes': uw_property_taxes,
                'insurance': uw_insurance,
                'utilities': uw_utilities,
                'maintenance_repairs': uw_maintenance_repairs,
                'management_fees': uw_management_fees,
                'replacement_reserves': uw_replacement_reserves,
                'total_expenses': uw_total_expenses,
                'expense_ratio': expense_ratio
            },
            'noi_analysis': {
                'net_operating_income': uw_net_operating_income,
                'cap_rate': market_cap_rate,
                'estimated_value': estimated_value,
                'price_per_unit': price_per_unit
            },
            'rent_analysis': rent_analysis,
            'actuals_vs_underwritten': {
                'actual_noi': t12_data.get('net_operating_income', 0),
                'underwritten_noi': uw_net_operating_income,
                'actual_vacancy_rate': actual_vacancy_rate,
                'underwritten_vacancy_rate': underwriting_vacancy_rate
            }
        }
    
    def _calculate_rm_minimum(self, total_units, property_age):
        """Calculate R&M minimum based on property age."""
        for (min_age, max_age), rate in self.config['rm_minimums'].items():
            if min_age <= property_age < max_age:
                return total_units * rate
        return total_units * 1000  # Default for very old properties
    
    def _calculate_management_fees(self, gross_potential_income):
        """Calculate management fees based on income tiers."""
        for min_income, max_income, rate in self.config['management_fee_tiers']:
            if min_income <= gross_potential_income < max_income:
                return gross_potential_income * rate
        return gross_potential_income * 0.03  # Default 3%
    
    def _generate_rent_analysis(self, rent_roll_df):
        """Generate detailed rent analysis."""
        
        occupied_units = rent_roll_df[rent_roll_df['Status'] == 'Occupied']
        vacant_units = rent_roll_df[rent_roll_df['Status'] == 'Vacant']
        
        rent_analysis = {
            'total_units': len(rent_roll_df),
            'occupied_units': len(occupied_units),
            'vacant_units': len(vacant_units),
            'occupancy_rate': len(occupied_units) / len(rent_roll_df) if len(rent_roll_df) > 0 else 0,
            'average_rent': occupied_units['Current_Rent'].mean() if len(occupied_units) > 0 else 0,
            'median_rent': occupied_units['Current_Rent'].median() if len(occupied_units) > 0 else 0,
            'min_rent': occupied_units['Current_Rent'].min() if len(occupied_units) > 0 else 0,
            'max_rent': occupied_units['Current_Rent'].max() if len(occupied_units) > 0 else 0,
            'rent_per_sqft': (occupied_units['Current_Rent'] / occupied_units['Square_Feet']).mean() if len(occupied_units) > 0 else 0
        }
        
        # Unit type breakdown
        unit_type_analysis = {}
        for unit_type in rent_roll_df['Unit_Type'].unique():
            type_units = rent_roll_df[rent_roll_df['Unit_Type'] == unit_type]
            occupied_type = type_units[type_units['Status'] == 'Occupied']
            
            unit_type_analysis[unit_type] = {
                'total_units': len(type_units),
                'occupied_units': len(occupied_type),
                'vacancy_rate': 1 - (len(occupied_type) / len(type_units)) if len(type_units) > 0 else 0,
                'average_rent': occupied_type['Current_Rent'].mean() if len(occupied_type) > 0 else 0,
                'average_sqft': type_units['Square_Feet'].mean(),
                'rent_per_sqft': (occupied_type['Current_Rent'] / occupied_type['Square_Feet']).mean() if len(occupied_type) > 0 else 0
            }
        
        rent_analysis['unit_type_breakdown'] = unit_type_analysis
        
        return rent_analysis
    
    def _create_professional_excel(self, analysis, rent_roll_data):
        """Create professional Excel package matching industry standards."""
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        data_font = Font(size=10)
        currency_format = '#,##0'
        percentage_format = '0.0%'
        
        # Create sheets
        self._create_property_characteristics_sheet(wb, analysis, header_font, header_fill, data_font)
        self._create_underwriting_summary_sheet(wb, analysis, header_font, header_fill, data_font)
        self._create_rent_roll_sheet(wb, rent_roll_data, analysis, header_font, header_fill, data_font)
        self._create_t12_analysis_sheet(wb, analysis, header_font, header_fill, data_font)
        
        # Save workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = f"outputs/Professional_Underwriting_Package_{timestamp}.xlsx"
        os.makedirs("outputs", exist_ok=True)
        wb.save(excel_path)
        
        return excel_path
    
    def _create_property_characteristics_sheet(self, wb, analysis, header_font, header_fill, data_font):
        """Create Property Characteristics sheet."""
        
        ws = wb.create_sheet("Property Characteristics")
        
        # Property Information Section
        ws['A1'] = "PROPERTY CHARACTERISTICS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        sections = [
            ("BASIC INFO", [
                ("Property Name", analysis['property_characteristics']['property_name']),
                ("Property Address", analysis['property_characteristics']['property_address']),
                ("Total Units", analysis['property_characteristics']['total_units']),
                ("Property Age", f"{analysis['property_characteristics']['property_age']} years"),
                ("Transaction Type", analysis['property_characteristics']['transaction_type'].title())
            ]),
            ("UNIT ANALYSIS", [
                ("Total Units", analysis['property_characteristics']['total_units']),
                ("Occupied Units", analysis['property_characteristics']['occupied_units']),
                ("Vacant Units", analysis['property_characteristics']['vacant_units']),
                ("Occupancy Rate", f"{(analysis['property_characteristics']['occupied_units'] / analysis['property_characteristics']['total_units']) * 100:.1f}%"),
                ("Average Rent", f"${analysis['property_characteristics']['average_rent']:,.0f}"),
                ("Average Sq Ft", f"{analysis['property_characteristics']['average_sqft']:,.0f}"),
                ("Rent per Sq Ft", f"${analysis['property_characteristics']['rent_per_sqft']:.2f}")
            ]),
            ("LOAN TERMS", [
                ("Loan Amount", f"${analysis['loan_terms']['loan_amount']:,.0f}"),
                ("Loan-to-Value", f"{analysis['loan_terms']['loan_to_value']:.1%}"),
                ("Interest Rate", f"{analysis['loan_terms']['interest_rate']:.3%}"),
                ("Amortization", f"{analysis['loan_terms']['amortization']} years"),
                ("Term", f"{analysis['loan_terms']['term']} years")
            ])
        ]
        
        for section_name, items in sections:
            # Section header
            ws[f'A{row}'] = section_name
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            # Section items
            for label, value in items:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = data_font
                ws[f'B{row}'].font = data_font
                row += 1
            
            row += 1  # Add space between sections
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
    
    def _create_underwriting_summary_sheet(self, wb, analysis, header_font, header_fill, data_font):
        """Create Underwriting Summary sheet matching the image format."""
        
        ws = wb.create_sheet("Underwriting Summary")
        
        # Title
        ws['A1'] = "UNDERWRITING SUMMARY"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = ['Line Item', 'Annual Amount', '% of EGI', 'Per Unit', 'Per Sq Ft', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 4
        egi = analysis['income_analysis']['effective_gross_income']
        total_units = analysis['property_characteristics']['total_units']
        total_sqft = total_units * analysis['property_characteristics']['average_sqft']
        
        # Income section
        income_items = [
            ("INCOME", "", "", "", "", ""),
            ("Gross Potential Income", analysis['income_analysis']['gross_potential_income'], 
             analysis['income_analysis']['gross_potential_income'] / egi, 
             analysis['income_analysis']['gross_potential_income'] / total_units,
             analysis['income_analysis']['gross_potential_income'] / total_sqft, ""),
            ("Less: Vacancy Loss", -analysis['income_analysis']['vacancy_loss'],
             -analysis['income_analysis']['vacancy_loss'] / egi,
             -analysis['income_analysis']['vacancy_loss'] / total_units,
             -analysis['income_analysis']['vacancy_loss'] / total_sqft,
             f"{analysis['income_analysis']['vacancy_rate']:.1%} vacancy rate"),
            ("Plus: Other Income", analysis['income_analysis']['other_income'],
             analysis['income_analysis']['other_income'] / egi,
             analysis['income_analysis']['other_income'] / total_units,
             analysis['income_analysis']['other_income'] / total_sqft, ""),
            ("Effective Gross Income", egi, 1.0, egi / total_units, egi / total_sqft, ""),
            ("", "", "", "", "", ""),
            ("OPERATING EXPENSES", "", "", "", "", ""),
            ("Property Taxes", analysis['expense_analysis']['property_taxes'],
             analysis['expense_analysis']['property_taxes'] / egi,
             analysis['expense_analysis']['property_taxes'] / total_units,
             analysis['expense_analysis']['property_taxes'] / total_sqft,
             "Adjusted for refinance"),
            ("Insurance", analysis['expense_analysis']['insurance'],
             analysis['expense_analysis']['insurance'] / egi,
             analysis['expense_analysis']['insurance'] / total_units,
             analysis['expense_analysis']['insurance'] / total_sqft, "Adjusted +5%"),
            ("Utilities", analysis['expense_analysis']['utilities'],
             analysis['expense_analysis']['utilities'] / egi,
             analysis['expense_analysis']['utilities'] / total_units,
             analysis['expense_analysis']['utilities'] / total_sqft, "Adjusted +2%"),
            ("Repairs & Maintenance", analysis['expense_analysis']['maintenance_repairs'],
             analysis['expense_analysis']['maintenance_repairs'] / egi,
             analysis['expense_analysis']['maintenance_repairs'] / total_units,
             analysis['expense_analysis']['maintenance_repairs'] / total_sqft,
             "Age-based minimum"),
            ("Management Fees", analysis['expense_analysis']['management_fees'],
             analysis['expense_analysis']['management_fees'] / egi,
             analysis['expense_analysis']['management_fees'] / total_units,
             analysis['expense_analysis']['management_fees'] / total_sqft,
             "Tier-based calculation"),
            ("Replacement Reserves", analysis['expense_analysis']['replacement_reserves'],
             analysis['expense_analysis']['replacement_reserves'] / egi,
             analysis['expense_analysis']['replacement_reserves'] / total_units,
             analysis['expense_analysis']['replacement_reserves'] / total_sqft,
             "$250/unit"),
            ("Total Operating Expenses", analysis['expense_analysis']['total_expenses'],
             analysis['expense_analysis']['expense_ratio'],
             analysis['expense_analysis']['total_expenses'] / total_units,
             analysis['expense_analysis']['total_expenses'] / total_sqft, ""),
            ("", "", "", "", "", ""),
            ("NET OPERATING INCOME", analysis['noi_analysis']['net_operating_income'],
             analysis['noi_analysis']['net_operating_income'] / egi,
             analysis['noi_analysis']['net_operating_income'] / total_units,
             analysis['noi_analysis']['net_operating_income'] / total_sqft, "")
        ]
        
        for item in income_items:
            for col, value in enumerate(item, 1):
                cell = ws.cell(row=row, column=col)
                
                if col == 1:  # Line Item
                    cell.value = value
                    if value in ["INCOME", "OPERATING EXPENSES", "NET OPERATING INCOME"]:
                        cell.font = Font(bold=True, size=11)
                    else:
                        cell.font = data_font
                elif col in [2, 4, 5] and isinstance(value, (int, float)) and value != "":  # Currency columns
                    cell.value = value
                    cell.number_format = '#,##0'
                    cell.font = data_font
                elif col == 3 and isinstance(value, (int, float)) and value != "":  # Percentage
                    cell.value = value
                    cell.number_format = '0.0%'
                    cell.font = data_font
                else:
                    cell.value = value
                    cell.font = data_font
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
    
    def _create_rent_roll_sheet(self, wb, rent_roll_data, analysis, header_font, header_fill, data_font):
        """Create Rent Roll Analysis sheet."""
        
        ws = wb.create_sheet("Rent Roll Analysis")
        
        # Title
        ws['A1'] = "RENT ROLL ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Summary section
        row = 3
        ws[f'A{row}'] = "RENT SUMMARY"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        rent_summary = [
            ("Total Units", analysis['rent_analysis']['total_units']),
            ("Occupied Units", analysis['rent_analysis']['occupied_units']),
            ("Vacant Units", analysis['rent_analysis']['vacant_units']),
            ("Occupancy Rate", f"{analysis['rent_analysis']['occupancy_rate']:.1%}"),
            ("Average Rent", f"${analysis['rent_analysis']['average_rent']:,.0f}"),
            ("Rent per Sq Ft", f"${analysis['rent_analysis']['rent_per_sqft']:.2f}")
        ]
        
        for label, value in rent_summary:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = data_font
            ws[f'B{row}'].font = data_font
            row += 1
        
        row += 2
        
        # Detailed rent roll
        ws[f'A{row}'] = "DETAILED RENT ROLL"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        # Headers
        headers = ['Unit #', 'Unit Type', 'Sq Ft', 'Current Rent', 'Status', 'Rent/SF', 'Annual Rent', 'Lease End']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        
        # Data
        for _, unit in rent_roll_data.iterrows():
            ws.cell(row=row, column=1, value=unit['Unit_Number'])
            ws.cell(row=row, column=2, value=unit['Unit_Type'])
            ws.cell(row=row, column=3, value=unit['Square_Feet'])
            ws.cell(row=row, column=4, value=unit['Current_Rent'])
            ws.cell(row=row, column=5, value=unit['Status'])
            ws.cell(row=row, column=6, value=unit['Current_Rent'] / unit['Square_Feet'] if unit['Square_Feet'] > 0 else 0)
            ws.cell(row=row, column=7, value=unit['Current_Rent'] * 12)
            ws.cell(row=row, column=8, value=unit.get('Lease_End_Date', ''))
            
            # Apply formatting
            for col in range(1, 9):
                ws.cell(row=row, column=col).font = data_font
            
            # Currency formatting
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=6).number_format = '0.00'
            ws.cell(row=row, column=7).number_format = '#,##0'
            
            row += 1
        
        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    def _create_t12_analysis_sheet(self, wb, analysis, header_font, header_fill, data_font):
        """Create T12 Analysis sheet."""
        
        ws = wb.create_sheet("T12 vs Underwritten")
        
        # Title
        ws['A1'] = "T12 ACTUAL vs UNDERWRITTEN COMPARISON"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ['Line Item', 'T12 Actual', 'Underwritten', 'Variance', 'Notes']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 4
        
        # Comparison data
        actual_noi = analysis['actuals_vs_underwritten']['actual_noi']
        uw_noi = analysis['actuals_vs_underwritten']['underwritten_noi']
        actual_vacancy = analysis['actuals_vs_underwritten']['actual_vacancy_rate']
        uw_vacancy = analysis['actuals_vs_underwritten']['underwritten_vacancy_rate']
        
        comparison_items = [
            ("Net Operating Income", actual_noi, uw_noi, uw_noi - actual_noi, "Underwritten with adjustments"),
            ("Vacancy Rate", actual_vacancy, uw_vacancy, uw_vacancy - actual_vacancy, "Minimum 5% applied"),
            ("Expense Ratio", "", analysis['expense_analysis']['expense_ratio'], "", "Includes all adjustments"),
            ("Cap Rate", "", analysis['noi_analysis']['cap_rate'], "", "Market rate applied"),
            ("Estimated Value", "", analysis['noi_analysis']['estimated_value'], "", "Based on underwritten NOI")
        ]
        
        for item in comparison_items:
            ws.cell(row=row, column=1, value=item[0]).font = data_font
            
            if isinstance(item[1], (int, float)) and item[1] != "":
                if "Rate" in item[0] or "Ratio" in item[0]:
                    ws.cell(row=row, column=2, value=item[1]).number_format = '0.0%'
                else:
                    ws.cell(row=row, column=2, value=item[1]).number_format = '#,##0'
            else:
                ws.cell(row=row, column=2, value=item[1])
            
            if isinstance(item[2], (int, float)) and item[2] != "":
                if "Rate" in item[0] or "Ratio" in item[0]:
                    ws.cell(row=row, column=3, value=item[2]).number_format = '0.0%'
                else:
                    ws.cell(row=row, column=3, value=item[2]).number_format = '#,##0'
            else:
                ws.cell(row=row, column=3, value=item[2])
            
            if isinstance(item[3], (int, float)) and item[3] != "":
                if "Rate" in item[0] or "Ratio" in item[0]:
                    ws.cell(row=row, column=4, value=item[3]).number_format = '0.0%'
                else:
                    ws.cell(row=row, column=4, value=item[3]).number_format = '#,##0'
            else:
                ws.cell(row=row, column=4, value=item[3])
            
            ws.cell(row=row, column=5, value=item[4]).font = data_font
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
    
    def _create_professional_pdf(self, analysis, excel_path):
        """Create professional PDF summary."""
        
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_path = f"outputs/Professional_Underwriting_Summary_{timestamp}.pdf"
        
        doc = SimpleDocTemplate(pdf_path, pagesize=letter, topMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=20)
        heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=12, spaceAfter=10)
        
        # Title
        story.append(Paragraph("PROFESSIONAL UNDERWRITING ANALYSIS", title_style))
        story.append(Paragraph(f"{analysis['property_characteristics']['property_name']}", styles['Heading1']))
        story.append(Paragraph(f"{analysis['property_characteristics']['property_address']}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        
        summary_data = [
            ['Property Type:', 'Multifamily'],
            ['Total Units:', f"{analysis['property_characteristics']['total_units']:,}"],
            ['Net Operating Income:', f"${analysis['noi_analysis']['net_operating_income']:,.0f}"],
            ['Estimated Value:', f"${analysis['noi_analysis']['estimated_value']:,.0f}"],
            ['Price per Unit:', f"${analysis['noi_analysis']['price_per_unit']:,.0f}"],
            ['Cap Rate:', f"{analysis['noi_analysis']['cap_rate']:.1%}"],
            ['Expense Ratio:', f"{analysis['expense_analysis']['expense_ratio']:.1%}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Financial Summary
        story.append(Paragraph("FINANCIAL SUMMARY", heading_style))
        
        financial_data = [
            ['', 'Annual Amount', '% of EGI'],
            ['Effective Gross Income', f"${analysis['income_analysis']['effective_gross_income']:,.0f}", '100.0%'],
            ['Total Operating Expenses', f"${analysis['expense_analysis']['total_expenses']:,.0f}", 
             f"{analysis['expense_analysis']['expense_ratio']:.1%}"],
            ['Net Operating Income', f"${analysis['noi_analysis']['net_operating_income']:,.0f}", 
             f"{analysis['noi_analysis']['net_operating_income'] / analysis['income_analysis']['effective_gross_income']:.1%}"]
        ]
        
        financial_table = Table(financial_data, colWidths=[2*inch, 1.5*inch, 1*inch])
        financial_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(financial_table)
        
        doc.build(story)
        return pdf_path

def main():
    """Main function to run the enhanced underwriting generator."""
    
    # Property information
    class PropertyInfo:
        def __init__(self):
            self.property_name = "Bolden Heights Apartments"
            self.property_address = "3350 Mount Gilead Road, Atlanta, GA 30311"
            self.transaction_type = "refinance"
            self.property_age = 25
            self.loan_amount = 15000000
    
    property_info = PropertyInfo()
    
    # File paths
    test_dir = "uploads/2ed8d504-4f6a-4bd2-ab3b-8cd5c137a5cb"
    rent_roll_path = f"{test_dir}/rent_roll/RR_3350_Mount_Gilead_Rd_Atlanta_GA_30311.pdf"
    t12_path = f"{test_dir}/t12/T12_3350_Mount_Gilead_Rd_Atlanta_GA_30311.pdf"
    
    # Generate package
    generator = EnhancedUnderwritingGenerator(debug=True)
    
    try:
        results = generator.generate_professional_package(rent_roll_path, t12_path, property_info)
        
        print(f"\n✅ Professional underwriting package generated successfully!")
        print(f"   📊 Excel: {results['excel_path']}")
        print(f"   📄 PDF: {results['pdf_path']}")
        print(f"   💰 NOI: ${results['analysis']['noi_analysis']['net_operating_income']:,.0f}")
        print(f"   📈 Estimated Value: ${results['analysis']['noi_analysis']['estimated_value']:,.0f}")
        print(f"   🏠 Units: {results['analysis']['property_characteristics']['total_units']}")
        
    except Exception as e:
        print(f"❌ Error generating package: {str(e)}")
        if generator.debug:
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()