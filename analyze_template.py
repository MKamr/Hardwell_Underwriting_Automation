#!/usr/bin/env python3
"""
Analyze the existing Excel template to understand its structure
"""

import pandas as pd
import openpyxl
import os

def analyze_excel_template():
    """Analyze the Excel template structure."""
    
    excel_path = "../Loan_Package_3350_Mount_Gilead_Rd_Atlanta_GA_30311_9_26_2024.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        return
    
    print("🔍 Analyzing Excel Template Structure...")
    print("=" * 80)
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        print(f"📊 Worksheet Names:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"   {i}. {sheet_name}")
        
        print("\n" + "=" * 80)
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n📋 Analyzing Sheet: '{sheet_name}'")
            print("-" * 50)
            
            ws = wb[sheet_name]
            
            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"   Dimensions: {max_row} rows × {max_col} columns")
            
            # Show first 20 rows with data
            print(f"   Sample Data (first 20 rows):")
            
            data_found = False
            for row in range(1, min(21, max_row + 1)):
                row_data = []
                has_data = False
                
                for col in range(1, min(10, max_col + 1)):  # First 10 columns
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    
                    if value is not None:
                        has_data = True
                        if isinstance(value, (int, float)):
                            if value > 1000:
                                row_data.append(f"${value:,.0f}")
                            else:
                                row_data.append(str(value))
                        else:
                            row_data.append(str(value)[:30])  # Truncate long strings
                    else:
                        row_data.append("")
                
                if has_data:
                    data_found = True
                    print(f"   Row {row:2d}: {' | '.join(row_data)}")
            
            if not data_found:
                print("   (No data found in first 20 rows)")
        
        print("\n" + "=" * 80)
        print("✅ Template analysis complete!")
        
        # Try to identify key sections
        print("\n🔍 Looking for key underwriting sections...")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Look for common underwriting terms
            underwriting_terms = [
                'INCOME', 'EXPENSE', 'NOI', 'Net Operating Income',
                'Rental Income', 'Other Income', 'Property Taxes',
                'Insurance', 'Management Fee', 'Vacancy', 'Utilities',
                'Repairs', 'Maintenance', 'Replacement Reserves'
            ]
            
            found_terms = []
            for row in range(1, min(100, ws.max_row + 1)):
                for col in range(1, min(20, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        for term in underwriting_terms:
                            if term.lower() in cell_value.lower():
                                found_terms.append(f"'{term}' found at {sheet_name}[{row},{col}]: {cell_value[:50]}")
            
            if found_terms:
                print(f"\n   📋 Sheet '{sheet_name}' - Found underwriting terms:")
                for term in found_terms[:10]:  # Show first 10 matches
                    print(f"     {term}")
        
    except Exception as e:
        print(f"❌ Error analyzing template: {str(e)}")

def create_template_mapper():
    """Create a template mapper based on the existing structure."""
    
    print("\n🔧 Creating Template Mapper...")
    print("=" * 80)
    
    # This will help us map our extracted data to the template
    template_mapping = {
        'underwriting_sheet': {
            'name': 'Underwriting',  # Common name for underwriting sheet
            'sections': {
                'income': {
                    'rental_income': 'Rental Income',
                    'other_income': 'Other Income', 
                    'gross_potential_income': 'Gross Potential Income',
                    'vacancy_loss': 'Vacancy Loss',
                    'effective_gross_income': 'Effective Gross Income'
                },
                'expenses': {
                    'property_taxes': 'Property Taxes',
                    'insurance': 'Insurance',
                    'utilities': 'Utilities',
                    'maintenance_repairs': 'Repairs & Maintenance',
                    'management_fees': 'Management Fees',
                    'replacement_reserves': 'Replacement Reserves',
                    'total_expenses': 'Total Operating Expenses'
                },
                'noi': {
                    'net_operating_income': 'Net Operating Income'
                }
            }
        },
        'rent_roll_sheet': {
            'name': 'Clean Rent Roll',
            'columns': [
                'Unit #', 'Unit Type', 'Sq Ft', 'Current Rent', 
                'Status', 'Lease End Date'
            ]
        },
        't12_sheet': {
            'name': 'Clean T12',
            'sections': [
                'Income Items', 'Operating Expenses', 'Net Operating Income'
            ]
        }
    }
    
    print("✅ Template mapping structure created!")
    return template_mapping

if __name__ == "__main__":
    analyze_excel_template()
    template_mapping = create_template_mapper()